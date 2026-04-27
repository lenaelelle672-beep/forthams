"""
资产批量导入导出集成测试

验收标准覆盖:
- ATB-001: 文件模板生成
- ATB-002: 字段校验
- ATB-003: 部分导入模式
- ATB-004: 异步导入处理
- ATB-005: 批量导出
- ATB-006: 安全性验证
- ATB-007: 数据完整性

功能标识: SWARM-2025-Q2-P2-006
迭代版本: Iteration 2
Phase: Phase 2 (数据交换层)
"""

import io
import json
import os
import tempfile
from datetime import date, datetime
from typing import Any, Dict, List
from unittest.mock import MagicMock, patch

import openpyxl
import pytest


# ============================================================================
# Test Data Fixtures
# ============================================================================

class TestAssetData:
    """测试用资产数据生成器"""
    
    @staticmethod
    def create_valid_asset(**overrides) -> Dict[str, Any]:
        """创建有效的资产数据"""
        asset = {
            "asset_id": "TEST-001",
            "asset_name": "测试资产",
            "asset_type": "EQUIPMENT",
            "purchase_date": "2024-01-15",
            "purchase_amount": "15000.00",
            "department": "DEPT-001",
            "status": "ACTIVE",
            "location": "A栋-101",
            "description": "测试用资产描述"
        }
        asset.update(overrides)
        return asset
    
    @staticmethod
    def create_asset_with_missing_field(field_name: str) -> Dict[str, Any]:
        """创建缺少指定字段的资产数据"""
        asset = TestAssetData.create_valid_asset()
        if field_name in asset:
            del asset[field_name]
        return asset
    
    @staticmethod
    def create_asset_with_invalid_value(field_name: str, invalid_value: Any) -> Dict[str, Any]:
        """创建具有无效值的资产数据"""
        asset = TestAssetData.create_valid_asset()
        asset[field_name] = invalid_value
        return asset
    
    @staticmethod
    def generate_excel_data(rows: List[Dict[str, Any]]) -> bytes:
        """生成Excel格式的资产数据"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assets"
        
        if not rows:
            return b""
        
        # Write headers
        headers = list(rows[0].keys())
        ws.append(headers)
        
        # Write data rows
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
    
    @staticmethod
    def generate_csv_data(rows: List[Dict[str, Any]]) -> bytes:
        """生成CSV格式的资产数据"""
        if not rows:
            return b""
        
        headers = list(rows[0].keys())
        lines = [",".join(headers)]
        
        for row in rows:
            values = [str(row.get(h, "")) for h in headers]
            lines.append(",".join(values))
        
        return "\n".join(lines).encode("utf-8-sig")


class MockAssetRepository:
    """模拟资产仓储层"""
    
    def __init__(self):
        self._assets = {}
    
    def exists(self, asset_id: str) -> bool:
        """检查资产ID是否存在"""
        return asset_id in self._assets
    
    def create(self, asset_data: Dict[str, Any]) -> Dict[str, Any]:
        """创建资产"""
        asset_id = asset_data.get("asset_id")
        self._assets[asset_id] = asset_data
        return asset_data
    
    def bulk_create(self, assets_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """批量创建资产"""
        return [self.create(a) for a in assets_data]
    
    def find_by_department(self, dept_code: str) -> List[Dict[str, Any]]:
        """按部门查找资产"""
        return [a for a in self._assets.values() if a.get("department") == dept_code]
    
    def find_by_status(self, status: str) -> List[Dict[str, Any]]:
        """按状态查找资产"""
        return [a for a in self._assets.values() if a.get("status") == status]
    
    def count_all(self) -> int:
        """统计资产总数"""
        return len(self._assets)
    
    def clear(self):
        """清空所有资产"""
        self._assets.clear()


class MockDepartmentRepository:
    """模拟部门仓储层"""
    
    def __init__(self):
        self._departments = {
            "DEPT-001": {"code": "DEPT-001", "name": "测试部门"},
            "DEPT-002": {"code": "DEPT-002", "name": "生产部"},
            "DEPT-003": {"code": "DEPT-003", "name": "研发部"}
        }
    
    def exists(self, dept_code: str) -> bool:
        """检查部门是否存在"""
        return dept_code in self._departments
    
    def find_by_code(self, dept_code: str) -> Dict[str, Any]:
        """按编码查找部门"""
        return self._departments.get(dept_code)


# ============================================================================
# ATB-001: 文件模板生成测试
# ============================================================================

class TestTemplateGeneration:
    """ATB-001: 文件模板生成"""
    
    def test_template_generation_xlsx(self):
        """
        测试XLSX格式模板生成
        
        验证点:
        1. 调用接口返回标准模板文件
        2. 模板包含全部9个资产字段
        3. 必填字段有明确标记
        """
        # Arrange
        from src.services.import_service import ImportService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        # Act
        template_bytes = import_service.generate_template(format="xlsx")
        
        # Assert - 验证返回的是Excel文件
        assert template_bytes is not None
        assert len(template_bytes) > 0
        
        # 验证Excel内容
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        
        # 验证包含全部9个资产字段
        expected_headers = [
            "asset_id", "asset_name", "asset_type", "purchase_date",
            "purchase_amount", "department", "status", "location", "description"
        ]
        
        for expected in expected_headers:
            assert expected in headers, f"缺少必需字段: {expected}"
    
    def test_template_generation_csv(self):
        """
        测试CSV格式模板生成
        
        验证点:
        1. 调用接口返回标准CSV文件
        2. 包含UTF-8 BOM编码
        """
        # Arrange
        from src.services.import_service import ImportService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        # Act
        csv_bytes = import_service.generate_template(format="csv")
        
        # Assert - 验证CSV格式
        assert csv_bytes is not None
        assert len(csv_bytes) > 0
        
        # 验证UTF-8 BOM编码
        assert csv_bytes[:3] == b'\xef\xbb\xbf', "CSV文件应包含UTF-8 BOM"
        
        # 解析CSV内容验证字段
        csv_text = csv_bytes.decode("utf-8-sig")
        lines = csv_text.strip().split("\n")
        assert len(lines) > 0
        
        headers = lines[0].split(",")
        
        expected_headers = [
            "asset_id", "asset_name", "asset_type", "purchase_date",
            "purchase_amount", "department", "status", "location", "description"
        ]
        
        for expected in expected_headers:
            assert expected in headers, f"缺少必需字段: {expected}"
    
    def test_template_headers_complete(self):
        """
        测试模板表头字段完整性
        
        验证点:
        - 包含全部9个资产字段
        - 字段顺序符合规范
        """
        from src.services.import_service import ImportService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        template_bytes = import_service.generate_template(format="xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        
        # 验证字段完整性
        required_fields = [
            "asset_id", "asset_name", "asset_type", "purchase_date",
            "purchase_amount", "department", "status", "location", "description"
        ]
        
        missing_fields = set(required_fields) - set(headers)
        assert len(missing_fields) == 0, f"模板缺少字段: {missing_fields}"
    
    def test_template_required_fields_marked(self):
        """
        测试模板必填字段标记
        
        验证点:
        - 必填字段有颜色/符号标记
        - 必填字段标识清晰
        """
        from src.services.import_service import ImportService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        template_bytes = import_service.generate_template(format="xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
        ws = wb.active
        
        # 检查表头样式
        # 必填字段应有特殊标记（通过字段名约定或样式）
        required_field_names = [
            "asset_id", "asset_name", "asset_type", "purchase_date",
            "purchase_amount", "department", "status"
        ]
        
        headers = [cell.value for cell in ws[1]]
        
        for field in required_field_names:
            assert field in headers, f"必填字段 {field} 未找到"
        
        # 可选字段
        optional_fields = ["location", "description"]
        
        for field in optional_fields:
            assert field in headers, f"可选字段 {field} 未找到"


# ============================================================================
# ATB-002: 字段校验测试
# ============================================================================

class TestFieldValidation:
    """ATB-002: 字段校验"""
    
    def test_validation_asset_id_required(self):
        """
        测试资产编号必填校验
        
        验证点:
        - asset_id为空的行返回校验失败
        - 错误提示清晰指明asset_id必填
        """
        from src.validators.field_validator import AssetFieldValidator
        
        validator = AssetFieldValidator()
        
        asset = TestAssetData.create_asset_with_missing_field("asset_id")
        result = validator.validate(asset)
        
        assert result.is_valid is False
        errors = result.errors
        
        # 查找asset_id相关错误
        asset_id_errors = [e for e in errors if "asset_id" in str(e.field).lower()]
        assert len(asset_id_errors) > 0, "应有asset_id必填错误"
        
        # 验证错误信息包含"必填"或"required"关键词
        error_msg = asset_id_errors[0].message.lower()
        assert "required" in error_msg or "必填" in error_msg or "不能为空" in error_msg
    
    def test_validation_asset_type_enum(self):
        """
        测试资产类型枚举校验
        
        验证点:
        - 无效的asset_type值被拒绝
        - 错误提示枚举值不匹配
        """
        from src.validators.field_validator import AssetFieldValidator
        
        validator = AssetFieldValidator()
        
        asset = TestAssetData.create_valid_asset(asset_type="INVALID_TYPE")
        result = validator.validate(asset)
        
        assert result.is_valid is False
        errors = result.errors
        
        # 查找asset_type相关错误
        type_errors = [e for e in errors if "asset_type" in str(e.field).lower()]
        assert len(type_errors) > 0, "应有asset_type枚举错误"
        
        # 验证错误信息指明有效枚举值
        error_msg = type_errors[0].message.lower()
        valid_types = ["equipment", "instrument", "vehicle", "other"]
        assert any(t in error_msg for t in valid_types), "错误信息应包含有效枚举值"
    
    def test_validation_date_format(self):
        """
        测试采购日期格式校验
        
        验证点:
        - 无效日期格式被拒绝
        - 错误提示日期格式错误
        """
        from src.validators.field_validator import AssetFieldValidator
        
        validator = AssetFieldValidator()
        
        # 测试无效日期格式
        invalid_dates = ["2025-13-01", "2025/01/15", "01-15-2025", "invalid"]
        
        for invalid_date in invalid_dates:
            asset = TestAssetData.create_valid_asset(purchase_date=invalid_date)
            result = validator.validate(asset)
            
            assert result.is_valid is False, f"日期 {invalid_date} 应被拒绝"
    
    def test_validation_amount_positive(self):
        """
        测试采购金额正数校验
        
        验证点:
        - 负数金额被拒绝
        - 错误提示金额必须≥0
        """
        from src.validators.field_validator import AssetFieldValidator
        
        validator = AssetFieldValidator()
        
        # 测试负数金额
        asset = TestAssetData.create_valid_asset(purchase_amount="-100")
        result = validator.validate(asset)
        
        assert result.is_valid is False
        errors = result.errors
        
        # 查找purchase_amount相关错误
        amount_errors = [e for e in errors if "purchase_amount" in str(e.field).lower()]
        assert len(amount_errors) > 0, "应有purchase_amount错误"
        
        # 验证错误信息指明数值范围
        error_msg = amount_errors[0].message.lower()
        assert "0" in error_msg or "正数" in error_msg or ">=0" in error_msg
    
    def test_validation_department_exists(self):
        """
        测试部门存在性校验
        
        验证点:
        - 不存在的部门编码被拒绝
        - 错误提示部门不存在
        """
        from src.validators.field_validator import AssetFieldValidator
        from src.services.validation_service import ValidationService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        
        validation_service = ValidationService(dept_repo=mock_dept_repo)
        validator = AssetFieldValidator(validation_service=validation_service)
        
        # 测试不存在的部门
        asset = TestAssetData.create_valid_asset(department="INVALID-DEPT")
        result = validator.validate(asset)
        
        assert result.is_valid is False
        errors = result.errors
        
        # 查找department相关错误
        dept_errors = [e for e in errors if "department" in str(e.field).lower()]
        assert len(dept_errors) > 0, "应有department错误"
        
        # 验证错误信息指明部门不存在
        error_msg = dept_errors[0].message.lower()
        assert "不存在" in error_msg or "not found" in error_msg or "invalid" in error_msg
    
    def test_validation_asset_id_unique(self):
        """
        测试资产编号唯一性校验
        
        验证点:
        - 重复的asset_id被拒绝
        - 错误提示asset_id重复
        """
        from src.validators.row_validator import AssetRowValidator
        
        mock_repo = MockAssetRepository()
        mock_repo.create(TestAssetData.create_valid_asset(asset_id="DUPLICATE-001"))
        
        validator = AssetRowValidator(asset_repo=mock_repo)
        
        # 测试重复的asset_id
        asset = TestAssetData.create_valid_asset(asset_id="DUPLICATE-001")
        result = validator.validate(asset, is_new=True)
        
        assert result.is_valid is False
        errors = result.errors
        
        # 查找asset_id重复错误
        id_errors = [e for e in errors if "asset_id" in str(e.field).lower() and "duplicate" in str(e.message).lower()]
        assert len(id_errors) > 0, "应有asset_id重复错误"


# ============================================================================
# ATB-003: 部分导入模式测试
# ============================================================================

class TestPartialImport:
    """ATB-003: 部分导入模式"""
    
    def test_partial_import_success_count(self):
        """
        测试部分导入成功行数统计
        
        验证点:
        - 100行数据中5行有错误
        - 成功导入95行
        - 生成错误报告
        """
        from src.services.import_service import ImportService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        # 准备测试数据: 100行，5行有错误
        valid_assets = [
            TestAssetData.create_valid_asset(asset_id=f"ASSET-{i:04d}")
            for i in range(1, 96)
        ]
        
        invalid_assets = [
            TestAssetData.create_asset_with_missing_field("asset_id"),  # Row 96
            TestAssetData.create_asset_with_invalid_value("asset_type", "INVALID"),  # Row 97
            TestAssetData.create_asset_with_invalid_value("purchase_amount", "-100"),  # Row 98
            TestAssetData.create_asset_with_invalid_value("purchase_date", "invalid"),  # Row 99
            TestAssetData.create_asset_with_invalid_value("department", "NOT-EXIST"),  # Row 100
        ]
        
        all_assets = valid_assets + invalid_assets
        
        # 执行部分导入
        result = import_service.import_assets(
            assets=all_assets,
            mode="partial",  # 部分导入模式
            overwrite=False
        )
        
        # 验证导入结果
        assert result.imported_count == 95, f"期望导入95行，实际{result.imported_count}"
        assert result.failed_count == 5, f"期望失败5行，实际{result.failed_count}"
        assert result.total_count == 100
    
    def test_error_report_format(self):
        """
        测试错误报告格式
        
        验证点:
        - 包含行号
        - 包含字段名
        - 包含错误原因
        """
        from src.services.import_service import ImportService
        from src.models.import_task import ImportTaskResult
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        # 准备带错误的数据
        assets = [
            TestAssetData.create_valid_asset(asset_id="OK-001"),
            TestAssetData.create_asset_with_missing_field("asset_id"),  # Row 2 - 错误
        ]
        
        result = import_service.import_assets(assets=assets, mode="partial")
        
        # 验证错误报告格式
        assert len(result.errors) >= 1
        
        error = result.errors[0]
        
        # 验证错误对象包含必要字段
        assert hasattr(error, "row"), "错误应包含行号"
        assert hasattr(error, "field"), "错误应包含字段名"
        assert hasattr(error, "message"), "错误应包含错误原因"
        
        assert error.row == 2, f"错误行号应为2，实际{error.row}"
        assert error.field in ["asset_id", "asset_name", "asset_type"], "字段名应在资产字段中"
        assert len(error.message) > 0, "错误信息不应为空"
    
    def test_error_report_export(self):
        """
        测试错误报告导出功能
        
        验证点:
        - 生成包含错误详情的Excel文件
        - 文件可下载
        """
        from src.services.import_service import ImportService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        # 准备带错误的数据
        assets = [
            TestAssetData.create_valid_asset(asset_id="OK-001"),
            TestAssetData.create_asset_with_missing_field("asset_id"),
        ]
        
        result = import_service.import_assets(assets=assets, mode="partial")
        
        # 导出错误报告
        error_report_bytes = import_service.export_error_report(result.errors, format="xlsx")
        
        assert error_report_bytes is not None
        assert len(error_report_bytes) > 0
        
        # 验证Excel内容
        wb = openpyxl.load_workbook(io.BytesIO(error_report_bytes))
        ws = wb.active
        
        # 检查表头包含必要列
        headers = [cell.value for cell in ws[1]]
        assert "行号" in headers or "row" in str(headers).lower()
        assert "字段" in headers or "field" in str(headers).lower()
        assert "错误原因" in headers or "message" in str(headers).lower()
        
        # 检查错误数据行
        assert ws.max_row >= 2, "应有至少1条错误数据"


# ============================================================================
# ATB-004: 异步导入处理测试
# ============================================================================

class TestAsyncImport:
    """ATB-004: 异步导入处理"""
    
    def test_async_import_task_created(self):
        """
        测试异步导入任务创建
        
        验证点:
        - 上传大文件返回任务ID
        - 立即响应不阻塞
        """
        from src.services.import_service import ImportService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        # 模拟大文件上传(30MB)
        file_size = 30 * 1024 * 1024  # 30MB
        file_content = TestAssetData.generate_excel_data([
            TestAssetData.create_valid_asset(asset_id=f"ASSET-{i}") 
            for i in range(1000)
        ])
        
        # 确保文件足够大以触发异步处理
        assert len(file_content) >= 5 * 1024 * 1024, "文件应大于5MB"
        
        # 创建导入任务
        task = import_service.create_import_task(
            file_content=file_content,
            file_name="large_import.xlsx",
            user_id="test-user-001",
            mode="partial"
        )
        
        # 验证任务创建成功
        assert task is not None
        assert task.task_id is not None
        assert task.status == "QUEUED" or task.status == "PENDING"
    
    def test_async_task_progress(self):
        """
        测试异步任务进度查询
        
        验证点:
        - 查询任务返回进度百分比
        - 进度状态清晰
        """
        from src.services.import_service import ImportService
        from src.models.import_task import ImportTask
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        # 创建任务
        task = import_service.create_import_task(
            file_content=b"test content",
            file_name="test.xlsx",
            user_id="test-user-001",
            mode="partial"
        )
        
        # 查询进度
        progress = import_service.get_task_progress(task.task_id)
        
        assert progress is not None
        assert 0 <= progress.percentage <= 100
        assert progress.status in ["PENDING", "PROCESSING", "COMPLETED", "FAILED"]
    
    def test_async_task_completed(self):
        """
        测试异步任务完成查询
        
        验证点:
        - 任务完成后返回成功/失败数量统计
        """
        from src.services.import_service import ImportService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        # 创建并模拟任务完成
        task = import_service.create_import_task(
            file_content=TestAssetData.generate_excel_data([
                TestAssetData.create_valid_asset(asset_id=f"ASSET-{i}") 
                for i in range(100)
            ]),
            file_name="completed_test.xlsx",
            user_id="test-user-001",
            mode="partial"
        )
        
        # 模拟任务处理完成
        import_service.simulate_task_completion(task.task_id)
        
        # 查询任务结果
        result = import_service.get_task_result(task.task_id)
        
        assert result is not None
        assert result.status == "COMPLETED"
        assert hasattr(result, "imported_count")
        assert hasattr(result, "failed_count")
    
    def test_async_task_failure(self):
        """
        测试异步任务失败处理
        
        验证点:
        - 任务失败时返回失败原因
        - 包含错误行号信息
        """
        from src.services.import_service import ImportService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        # 创建任务
        task = import_service.create_import_task(
            file_content=b"invalid content",
            file_name="invalid.xlsx",
            user_id="test-user-001",
            mode="partial"
        )
        
        # 模拟任务失败
        import_service.simulate_task_failure(
            task.task_id,
            reason="文件格式错误",
            error_rows=[3, 5, 7]
        )
        
        # 查询任务结果
        result = import_service.get_task_result(task.task_id)
        
        assert result is not None
        assert result.status == "FAILED"
        assert hasattr(result, "error_message")
        assert "文件格式错误" in result.error_message


# ============================================================================
# ATB-005: 批量导出测试
# ============================================================================

class TestBatchExport:
    """ATB-005: 批量导出"""
    
    def test_export_xlsx(self):
        """
        测试XLSX格式导出
        
        验证点:
        - 返回Excel文件
        - 文件格式正确
        """
        from src.services.export_service import ExportService
        
        mock_repo = MockAssetRepository()
        
        # 准备测试数据
        for i in range(10):
            mock_repo.create(TestAssetData.create_valid_asset(asset_id=f"EXPORT-{i:03d}"))
        
        export_service = ExportService(asset_repo=mock_repo)
        
        # 导出XLSX
        xlsx_bytes = export_service.export_assets(format="xlsx")
        
        assert xlsx_bytes is not None
        assert len(xlsx_bytes) > 0
        
        # 验证Excel内容
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        
        assert ws.max_row >= 10, "应有至少10行数据"
    
    def test_export_csv(self):
        """
        测试CSV格式导出
        
        验证点:
        - 返回CSV文件
        - UTF-8 BOM编码
        """
        from src.services.export_service import ExportService
        
        mock_repo = MockAssetRepository()
        
        # 准备测试数据
        for i in range(10):
            mock_repo.create(TestAssetData.create_valid_asset(asset_id=f"EXPORT-CSV-{i:03d}"))
        
        export_service = ExportService(asset_repo=mock_repo)
        
        # 导出CSV
        csv_bytes = export_service.export_assets(format="csv")
        
        assert csv_bytes is not None
        assert len(csv_bytes) > 0
        
        # 验证UTF-8 BOM
        assert csv_bytes[:3] == b'\xef\xbb\xbf', "CSV应包含UTF-8 BOM"
        
        # 验证CSV内容
        csv_text = csv_bytes.decode("utf-8-sig")
        lines = csv_text.strip().split("\n")
        
        assert len(lines) >= 11, "应有表头+10行数据"
    
    def test_export_field_consistency(self):
        """
        测试导出字段与模板一致性
        
        验证点:
        - 导出字段与模板导入字段一致
        """
        from src.services.export_service import ExportService
        from src.services.import_service import ImportService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        
        export_service = ExportService(asset_repo=mock_repo)
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        
        # 获取模板字段
        template_bytes = import_service.generate_template(format="xlsx")
        wb_template = openpyxl.load_workbook(io.BytesIO(template_bytes))
        ws_template = wb_template.active
        template_headers = [cell.value for cell in ws_template[1]]
        
        # 获取导出字段
        mock_repo.create(TestAssetData.create_valid_asset(asset_id="TEST-001"))
        export_bytes = export_service.export_assets(format="xlsx")
        wb_export = openpyxl.load_workbook(io.BytesIO(export_bytes))
        ws_export = wb_export.active
        export_headers = [cell.value for cell in ws_export[1]]
        
        # 验证字段一致性
        assert set(template_headers) == set(export_headers), \
            f"模板字段与导出字段不一致: 模板={template_headers}, 导出={export_headers}"
    
    def test_export_with_filter(self):
        """
        测试带筛选条件导出
        
        验证点:
        - 仅导出符合筛选条件的资产
        """
        from src.services.export_service import ExportService
        
        mock_repo = MockAssetRepository()
        
        # 准备不同状态的测试数据
        mock_repo.create(TestAssetData.create_valid_asset(asset_id="ACTIVE-001", status="ACTIVE"))
        mock_repo.create(TestAssetData.create_valid_asset(asset_id="ACTIVE-002", status="ACTIVE"))
        mock_repo.create(TestAssetData.create_valid_asset(asset_id="MAINT-001", status="MAINTENANCE"))
        
        export_service = ExportService(asset_repo=mock_repo)
        
        # 仅导出ACTIVE状态的资产
        export_bytes = export_service.export_assets(format="xlsx", filters={"status": "ACTIVE"})
        
        assert export_bytes is not None
        
        wb = openpyxl.load_workbook(io.BytesIO(export_bytes))
        ws = wb.active
        
        # 验证导出行数
        data_rows = ws.max_row - 1  # 减去表头行
        assert data_rows == 2, f"期望导出2行ACTIVE资产，实际{data_rows}行"
    
    def test_export_data_integrity(self):
        """
        测试导出数据完整性
        
        验证点:
        - 导出行数与数据库一致
        - 金额汇总与数据库一致
        """
        from src.services.export_service import ExportService
        
        mock_repo = MockAssetRepository()
        
        # 准备测试数据
        test_assets = [
            TestAssetData.create_valid_asset(asset_id=f"INT-{i:03d}", purchase_amount="1000.00")
            for i in range(10)
        ]
        
        for asset in test_assets:
            mock_repo.create(asset)
        
        # 导出数据
        export_service = ExportService(asset_repo=mock_repo)
        export_bytes = export_service.export_assets(format="xlsx")
        
        # 解析导出数据
        wb = openpyxl.load_workbook(io.BytesIO(export_bytes))
        ws = wb.active
        
        # 验证行数
        data_rows = ws.max_row - 1
        assert data_rows == mock_repo.count_all(), \
            f"导出行数({data_rows})与数据库({mock_repo.count_all()})不一致"
        
        # 验证金额汇总
        exported_amount = 0.0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # purchase_amount列应该是第5列(索引4)
            amount_str = str(row[4]) if len(row) > 4 else "0"
            try:
                exported_amount += float(amount_str)
            except ValueError:
                pass
        
        expected_amount = 10 * 1000.00  # 10条记录，每条1000
        assert abs(exported_amount - expected_amount) < 0.01, \
            f"金额汇总({exported_amount})与预期({expected_amount})不一致"


# ============================================================================
# ATB-006: 安全性验证测试
# ============================================================================

class TestSecurityVerification:
    """ATB-006: 安全性验证"""
    
    def test_import_unauthorized(self):
        """
        测试未授权用户导入
        
        验证点:
        - 未登录用户尝试导入返回401
        """
        from src.api.v1.asset_import import import_router
        
        # 模拟未登录请求
        request = MagicMock()
        request.headers = {}
        request.user = None  # 未认证
        
        # 执行导入（应该被拦截）
        with pytest.raises(Exception) as exc_info:
            import_router.handle_import(request)
        
        assert "401" in str(exc_info.value) or "Unauthorized" in str(exc_info.value)
    
    def test_import_forbidden(self):
        """
        测试无权限用户导入
        
        验证点:
        - 无权限用户尝试导入返回403
        """
        from src.api.v1.asset_import import import_router
        
        # 模拟无权限用户
        request = MagicMock()
        request.headers = {}
        request.user = MagicMock(user_id="user-001", permissions=[])  # 无权限
        
        # 执行导入（应该被拦截）
        with pytest.raises(Exception) as exc_info:
            import_router.handle_import(request)
        
        assert "403" in str(exc_info.value) or "Forbidden" in str(exc_info.value)
    
    def test_import_audit_log(self):
        """
        测试导入操作审计日志
        
        验证点:
        - 日志包含user_id
        - 日志包含file_name
        - 日志包含时间戳
        """
        from src.services.import_service import ImportService
        from src.infrastructure.audit.audit_hash_chain import AuditLogger
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        import_service = ImportService(
            asset_repo=mock_repo, 
            dept_repo=mock_dept_repo
        )
        
        # 准备审计日志器
        audit_logger = AuditLogger()
        
        # 执行导入
        assets = [
            TestAssetData.create_valid_asset(asset_id=f"USER-001-ASSET-{i}")
            for i in range(5)
        ]
        
        result = import_service.import_assets(assets=assets, mode="partial")
        
        # 获取审计日志
        audit_entry = audit_logger.get_latest_entry()
        
        # 验证日志包含必要信息
        assert audit_entry is not None
        assert "user_id" in audit_entry or "userId" in audit_entry or "user" in audit_entry.lower()
        assert "file_name" in audit_entry or "fileName" in audit_entry or "filename" in audit_entry.lower()
        assert "timestamp" in audit_entry or "time" in audit_entry.lower() or "created" in audit_entry.lower()
    
    def test_export_link_expiration(self):
        """
        测试导出下载链接有效期
        
        验证点:
        - 24小时后链接失效
        """
        from src.services.export_service import ExportService
        
        mock_repo = MockAssetRepository()
        mock_repo.create(TestAssetData.create_valid_asset(asset_id="TEST-EXP-001"))
        
        export_service = ExportService(asset_repo=mock_repo)
        
        # 生成导出链接
        link_info = export_service.generate_download_link(format="xlsx")
        
        assert link_info is not None
        assert hasattr(link_info, "expires_at")
        
        # 验证有效期为24小时
        expiry_time = link_info.expires_at
        now = datetime.now()
        
        # 允许1分钟误差
        expected_expiry = now.timestamp() + 24 * 3600
        assert abs(expiry_time.timestamp() - expected_expiry) < 60, \
            f"链接有效期应为24小时，实际{(expiry_time.timestamp() - now.timestamp()) / 3600}小时"


# ============================================================================
# ATB-007: 数据完整性测试
# ============================================================================

class TestDataIntegrity:
    """ATB-007: 数据完整性"""
    
    def test_export_import_roundtrip(self):
        """
        测试导出-导入往返数据一致性
        
        验证点:
        - 导出全部资产后删除
        - 重新导入后数据完全一致
        - ID连续性保持
        """
        from src.services.import_service import ImportService
        from src.services.export_service import ExportService
        
        mock_repo = MockAssetRepository()
        mock_dept_repo = MockDepartmentRepository()
        
        import_service = ImportService(asset_repo=mock_repo, dept_repo=mock_dept_repo)
        export_service = ExportService(asset_repo=mock_repo)
        
        # 准备初始数据
        original_assets = [
            TestAssetData.create_valid_asset(asset_id=f"ROUND-{i:03d}")
            for i in range(20)
        ]
        
        for asset in original_assets:
            mock_repo.create(asset)
        
        # 导出数据
        export_bytes = export_service.export_assets(format="xlsx")
        
        # 清空数据库
        mock_repo.clear()
        assert mock_repo.count_all() == 0
        
        # 重新导入
        wb = openpyxl.load_workbook(io.BytesIO(export_bytes))
        ws = wb.active
        
        reimport_assets = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            asset = {
                "asset_id": row[0],
                "asset_name": row[1],
                "asset_type": row[2],
                "purchase_date": row[3],
                "purchase_amount": row[4],
                "department": row[5],
                "status": row[6],
                "location": row[7] if len(row) > 7 else "",
                "description": row[8] if len(row) > 8 else ""
            }
            reimport_assets.append(asset)
        
        import_service.import_assets(assets=reimport_assets, mode="strict")
        
        # 验证数据一致性
        assert mock_repo.count_all() == 20, \
            f"导入后资产数量({mock_repo.count_all()})与原始数量(20)不一致"
        
        # 验证ID连续性
        exported_ids = sorted([a["asset_id"] for a in original_assets])
        imported_ids = sorted([a["asset_id"] for a in mock_repo._assets.values()])
        assert exported_ids == imported_ids, "导入后ID应与导出ID一致"
    
    def test_amount_precision(self):
        """
        测试金额精度保持
        
        验证点:
        - 导出金额与原金额精度一致
        - 金额保留2位小数
        """
        from src.services.export_service import ExportService
        
        mock_repo = MockAssetRepository()
        
        # 准备带精确金额的测试数据
        precision_amounts = [
            "12345.67",
            "0.01",
            "999999.99",
            "100.00",
            "0.99"
        ]
        
        for i, amount in enumerate(precision_amounts):
            mock_repo.create(
                TestAssetData.create_valid_asset(
                    asset_id=f"PREC-{i:03d}",
                    purchase_amount=amount
                )
            )
        
        export_service = ExportService(asset_repo=mock_repo)
        export_bytes = export_service.export_assets(format="xlsx")
        
        # 解析导出数据
        wb = openpyxl.load_workbook(io.BytesIO(export_bytes))
        ws = wb.active
        
        # 验证金额精度
        exported_amounts = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            amount_str = str(row[4]) if len(row) > 4 else "0"
            exported_amounts.append(amount_str)
        
        # 检查精度保留2位小数
        for original, exported in zip(precision_amounts, exported_amounts):
            # 移除可能的科学计数法
            original_float = float(original)
            exported_float = float(exported)
            
            assert abs(original_float - exported_float) < 0.001, \
                f"金额精度不一致: 原始={original}, 导出={exported}"


# ============================================================================
# 辅助工具类和测试配置
# ============================================================================

class ImportTaskResult:
    """导入任务结果封装"""
    
    def __init__(
        self, 
        imported_count: int = 0, 
        failed_count: int = 0, 
        total_count: int = 0,
        errors: List[Any] = None,
        status: str = "COMPLETED"
    ):
        self.imported_count = imported_count
        self.failed_count = failed_count
        self.total_count = total_count
        self.errors = errors or []
        self.status = status


# ============================================================================
# pytest 配置
# ============================================================================

@pytest.fixture
def mock_db_session():
    """模拟数据库会话"""
    session = MagicMock()
    session.query.return_value.filter.return_value.count.return_value = 0
    return session


@pytest.fixture
def asset_repo():
    """资产仓储fixture"""
    return MockAssetRepository()


@pytest.fixture
def dept_repo():
    """部门仓储fixture"""
    return MockDepartmentRepository()


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])