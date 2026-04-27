"""
Asset Batch Import/Export Test Suite
SWARM-2025-Q2-P2-006: 资产批量导入导出功能测试

本测试模块验证批量导入导出功能的完整性和正确性。
"""

import pytest
import io
import csv
from unittest.mock import Mock, patch, MagicMock
from datetime import datetime, date
from decimal import Decimal


# =============================================================================
# 测试数据 Fixtures
# =============================================================================

@pytest.fixture
def valid_asset_row():
    """
    生成有效资产数据行。
    
    Returns:
        dict: 符合字段映射表的资产数据
    """
    return {
        "asset_name": "测试资产-笔记本",
        "asset_type": "IT_HARDWARE",
        "serial_number": "SN-2025-001",
        "purchase_date": "2025-01-15",
        "purchase_price": "5999.00",
        "currency": "CNY",
        "department": "IT-DEPT",
        "custodian": "张三",
        "status": "ACTIVE",
        "location": "北京总部-5层",
        "remarks": "2025年采购"
    }


@pytest.fixture
def valid_csv_content(valid_asset_row):
    """
    生成标准 CSV 文件内容。
    
    Args:
        valid_asset_row: 有效资产数据行 fixture
    
    Returns:
        str: CSV 格式字符串
    """
    header = ",".join(valid_asset_row.keys())
    values = ",".join(str(v) for v in valid_asset_row.values())
    return f"{header}\n{values}"


@pytest.fixture
def invalid_asset_row():
    """
    生成包含校验错误的数据行。
    
    Returns:
        dict: 包含必填字段缺失和枚举值错误的数据
    """
    return {
        "asset_name": "",  # 必填字段为空
        "asset_type": "INVALID_TYPE",  # 非法枚举值
        "serial_number": "SN-001",
        "purchase_date": "2025-13-45",  # 非法日期格式
        "purchase_price": "-100",  # 价格必须大于0
        "currency": "CNY",
        "department": "IT-DEPT",
        "custodian": "李四",
        "status": "ACTIVE",
        "location": "上海分部",
        "remarks": ""
    }


@pytest.fixture
def mock_asset_service():
    """
    创建模拟的资产服务。
    
    Returns:
        Mock: 可配置返回值的资产服务模拟对象
    """
    service = Mock()
    service.create_asset = Mock(return_value={"id": "asset-001", "success": True})
    service.batch_create = Mock(return_value={"created": 100, "failed": 0})
    service.export_assets = Mock(return_value=[])
    return service


@pytest.fixture
def mock_department_service():
    """
    创建模拟的部门服务。
    
    Returns:
        Mock: 部门数据模拟对象
    """
    service = Mock()
    service.validate_department = Mock(return_value=True)
    service.get_by_code = Mock(return_value={"code": "IT-DEPT", "name": "IT部门"})
    return service


# =============================================================================
# TC-1: CSV 文件解析测试 (ATB-1.1)
# =============================================================================

class TestCsvParser:
    """
    CSV 文件解析器测试类。
    验证 CSV 文件的解析、编码处理和异常场景。
    """
    
    def test_parse_valid_csv(self, valid_csv_content):
        """
        TC-1.1.1: 正常 CSV 解析验证。
        
        测试标准 CSV 文件能够被正确解析为数据结构。
        """
        # Arrange
        from src.services.parsers.csv_parser import CsvParser
        
        # Act
        parser = CsvParser()
        result = parser.parse(io.StringIO(valid_csv_content))
        
        # Assert
        assert result is not None
        assert len(result) == 1
        assert result[0]["asset_name"] == "测试资产-笔记本"
        assert result[0]["asset_type"] == "IT_HARDWARE"
    
    def test_parse_utf8_with_bom(self):
        """
        TC-1.1.2: UTF-8-BOM 编码文件解析。
        
        验证带 BOM 的 UTF-8 文件能够正确解析。
        """
        # Arrange
        from src.services.parsers.csv_parser import CsvParser
        
        csv_content = "\ufeffasset_name,asset_type\n测试资产,IT_HARDWARE"
        
        # Act
        parser = CsvParser()
        result = parser.parse(io.StringIO(csv_content))
        
        # Assert
        assert len(result) == 1
        assert result[0]["asset_name"] == "测试资产"
    
    def test_parse_gbk_encoding(self):
        """
        TC-1.1.3: GBK 编码兼容性测试。
        
        验证 GBK 编码的 CSV 文件能够正确解析。
        """
        # Arrange
        from src.services.parsers.csv_parser import CsvParser
        
        csv_bytes = "资产名称,资产类型\n笔记本电脑,IT_HARDWARE".encode("gbk")
        csv_stream = io.BytesIO(csv_bytes)
        
        # Act
        parser = CsvParser()
        result = parser.parse(csv_stream, encoding="gbk")
        
        # Assert
        assert len(result) == 1
        assert "笔记本电脑" in result[0]["资产名称"]
    
    def test_parse_with_empty_rows(self):
        """
        TC-1.1.4: 异常行处理测试。
        
        验证 CSV 中的空行能够被正确跳过并记录警告。
        """
        # Arrange
        from src.services.parsers.csv_parser import CsvParser
        
        csv_content = "asset_name,asset_type\n资产A,IT_HARDWARE\n\n资产B,FURNITURE\n"
        
        # Act
        parser = CsvParser()
        result = parser.parse(io.StringIO(csv_content))
        warnings = parser.get_warnings()
        
        # Assert
        assert len(result) == 2
        assert any("empty row" in w.lower() for w in warnings)
    
    def test_reject_file_over_10mb(self):
        """
        TC-1.1.5: 文件大小限制测试。
        
        验证超过 10MB 的文件被正确拒绝。
        """
        # Arrange
        from src.services.parsers.csv_parser import CsvParser
        from src.models.exceptions import FileSizeExceededError
        
        # 模拟 11MB 文件
        large_content = "x" * (11 * 1024 * 1024)
        
        # Act & Assert
        parser = CsvParser()
        with pytest.raises(FileSizeExceededError) as exc_info:
            parser.parse(io.StringIO(large_content))
        
        assert "10MB" in str(exc_info.value)


# =============================================================================
# TC-2: 字段校验逻辑测试 (ATB-1.2)
# =============================================================================

class TestFieldValidator:
    """
    字段校验器测试类。
    验证必填字段、枚举值、日期格式、数值格式等校验逻辑。
    """
    
    def test_missing_required_fields(self, valid_asset_row):
        """
        TC-1.2.1: 必填字段缺失验证。
        
        验证当必填字段为空时，能够正确识别并报告错误位置。
        """
        # Arrange
        from src.services.validators.asset_validator import AssetFieldValidator
        
        test_row = valid_asset_row.copy()
        test_row["asset_name"] = ""  # 必填字段设为空
        
        # Act
        validator = AssetFieldValidator()
        errors = validator.validate(test_row)
        
        # Assert
        assert len(errors) > 0
        error_fields = [e["field"] for e in errors]
        assert "asset_name" in error_fields
    
    def test_invalid_enum_value(self, valid_asset_row):
        """
        TC-1.2.2: 枚举值越界验证。
        
        验证非法枚举值能够被正确识别。
        """
        # Arrange
        from src.services.validators.asset_validator import AssetFieldValidator
        
        test_row = valid_asset_row.copy()
        test_row["asset_type"] = "INVALID_ENUM"
        test_row["status"] = "INVALID_STATUS"
        
        # Act
        validator = AssetFieldValidator()
        errors = validator.validate(test_row)
        
        # Assert
        error_dict = {e["field"]: e for e in errors}
        assert "asset_type" in error_dict
        assert error_dict["asset_type"]["reason"] == "invalid enum"
        assert "INVALID_ENUM" in error_dict["asset_type"]["value"]
    
    def test_invalid_date_format(self, valid_asset_row):
        """
        TC-1.2.3: 日期格式错误验证。
        
        验证非标准日期格式能够被正确拒绝。
        """
        # Arrange
        from src.services.validators.asset_validator import AssetFieldValidator
        
        test_row = valid_asset_row.copy()
        test_row["purchase_date"] = "2025/01/15"  # 错误格式，应为 YYYY-MM-DD
        
        # Act
        validator = AssetFieldValidator()
        errors = validator.validate(test_row)
        
        # Assert
        date_errors = [e for e in errors if e["field"] == "purchase_date"]
        assert len(date_errors) > 0
        assert "format" in date_errors[0]["reason"].lower()
    
    def test_non_numeric_price(self, valid_asset_row):
        """
        TC-1.2.4: 价格非数字验证。
        
        验证非数字价格值能够被正确识别。
        """
        # Arrange
        from src.services.validators.asset_validator import AssetFieldValidator
        
        test_row = valid_asset_row.copy()
        test_row["purchase_price"] = "not-a-number"
        
        # Act
        validator = AssetFieldValidator()
        errors = validator.validate(test_row)
        
        # Assert
        price_errors = [e for e in errors if e["field"] == "purchase_price"]
        assert len(price_errors) > 0
        assert "must be numeric" in price_errors[0]["error"].lower()


# =============================================================================
# TC-3: 批量入库验证测试 (ATB-1.3)
# =============================================================================

class TestBatchImport:
    """
    批量导入业务逻辑测试类。
    验证同步/异步入库、部分失败回滚等场景。
    """
    
    def test_sync_import_500_rows(self, mock_asset_service, valid_asset_row):
        """
        TC-1.3.1: 500条同步入库测试。
        
        验证小批量数据能够快速同步完成。
        """
        # Arrange
        from src.application.services.batch_import_service import BatchImportService
        
        # 生成500条测试数据
        test_data = [valid_asset_row.copy() for _ in range(500)]
        for i, row in enumerate(test_data):
            row["asset_name"] = f"测试资产-{i:04d}"
        
        # Act
        service = BatchImportService(asset_service=mock_asset_service)
        start_time = datetime.now()
        result = service.import_sync(test_data)
        elapsed = (datetime.now() - start_time).total_seconds()
        
        # Assert
        assert result["success"] is True
        assert result["imported"] == 500
        assert elapsed < 5.0  # 应在5秒内完成
    
    def test_async_import_3000_rows(self, mock_asset_service, valid_asset_row):
        """
        TC-1.3.2: 3000条异步入库测试。
        
        验证大批量数据触发异步处理并返回任务ID。
        """
        # Arrange
        from src.application.services.batch_import_service import BatchImportService
        
        test_data = [valid_asset_row.copy() for _ in range(3000)]
        
        # Act
        service = BatchImportService(asset_service=mock_asset_service)
        result = service.import_async(test_data)
        
        # Assert
        assert "task_id" in result
        assert result["status"] == "PENDING"
        assert result["estimated_time"] > 0
    
    def test_partial_failure_rollback(self, mock_asset_service, valid_asset_row):
        """
        TC-1.3.3: 部分失败回滚测试。
        
        验证任意一条失败时全量回滚。
        """
        # Arrange
        from src.application.services.batch_import_service import BatchImportService
        
        test_data = [valid_asset_row.copy() for _ in range(10)]
        test_data[5]["asset_name"] = ""  # 注入错误
        
        # 配置服务在特定条件下抛出异常
        mock_asset_service.batch_create.side_effect = Exception("Database error")
        
        # Act
        service = BatchImportService(asset_service=mock_asset_service)
        result = service.import_sync(test_data)
        
        # Assert
        assert result["success"] is False
        assert result["status"] == "FAILED"
        assert result["rolled_back"] is True
    
    def test_import_task_status_query(self, mock_asset_service):
        """
        TC-1.4.2: 导入任务状态查询测试。
        
        验证能够通过 task_id 查询任务状态。
        """
        # Arrange
        from src.application.services.batch_import_service import BatchImportService
        
        service = BatchImportService(asset_service=mock_asset_service)
        task_id = "task-2025-001"
        
        # Act
        status = service.get_task_status(task_id)
        
        # Assert
        assert status is not None
        assert "status" in status
        assert "progress" in status


# =============================================================================
# TC-4: 错误报告生成测试 (ATB-1.4)
# =============================================================================

class TestImportReport:
    """
    导入错误报告测试类。
    验证错误报告格式和下载接口。
    """
    
    def test_error_report_csv_format(self, invalid_asset_row):
        """
        TC-1.4.1: 错误报告 CSV 格式测试。
        
        验证生成的错误报告包含正确的字段。
        """
        # Arrange
        from src.application.services.import_report_service import ImportReportService
        
        errors = [
            {"row": 5, "field": "asset_name", "value": "", "reason": "required field"},
            {"row": 5, "field": "asset_type", "value": "INVALID", "reason": "invalid enum"},
        ]
        
        # Act
        service = ImportReportService()
        report = service.generate_error_csv(errors)
        
        # Assert
        assert "row_number" in report
        assert "error_field" in report
        assert "error_detail" in report
    
    def test_error_report_download(self):
        """
        TC-1.4.2: 错误报告下载接口测试。
        
        验证错误报告能够正确生成并返回下载流。
        """
        # Arrange
        from src.application.services.import_report_service import ImportReportService
        from fastapi.testclient import TestClient
        from fastapi import FastAPI
        
        app = FastAPI()
        
        @app.get("/api/v1/asset-import/tasks/{task_id}/report")
        async def download_report(task_id: str):
            service = ImportReportService()
            return service.create_download_response(task_id)
        
        # Act
        client = TestClient(app)
        response = client.get("/api/v1/asset-import/tasks/task-001/report")
        
        # Assert
        assert response.status_code == 200
        assert response.headers["content-type"] == "text/csv"


# =============================================================================
# TC-5: Excel 导入功能测试 (ATB-2)
# =============================================================================

class TestExcelParser:
    """
    Excel 文件解析器测试类。
    验证 Excel 文件的解析和多工作表处理。
    """
    
    def test_parse_xlsx_worksheet(self):
        """
        TC-2.1: Excel 文件解析测试。
        
        验证 .xlsx 文件能够被正确解析。
        """
        # Arrange
        from src.services.parsers.excel_parser import ExcelParser
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.append(["asset_name", "asset_type"])
        ws.append(["测试资产", "IT_HARDWARE"])
        
        # Act
        parser = ExcelParser()
        result = parser.parse_workbook(wb)
        
        # Assert
        assert len(result) == 1
        assert result[0]["asset_name"] == "测试资产"
    
    def test_ignore_other_worksheets(self):
        """
        TC-2.2: 多工作表处理测试。
        
        验证非首工作表数据被正确忽略。
        """
        # Arrange
        from src.services.parsers.excel_parser import ExcelParser
        from openpyxl import Workbook
        
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet"
        ws1.append(["asset_name"], ["资产A"])
        
        ws2 = wb.create_sheet("Other")
        ws2.append(["asset_name"], ["不应包含的数据"])
        
        # Act
        parser = ExcelParser()
        result = parser.parse_workbook(wb)
        
        # Assert
        assert len(result) == 1
        assert result[0]["asset_name"] == "资产A"
    
    def test_empty_cell_as_empty_string(self):
        """
        TC-2.3: 空单元格处理测试。
        
        验证空单元格被转换为空字符串。
        """
        # Arrange
        from src.services.parsers.excel_parser import ExcelParser
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.append(["asset_name", "remarks"])
        ws.append(["资产A", None])  # 空单元格
        
        # Act
        parser = ExcelParser()
        result = parser.parse_workbook(wb)
        
        # Assert
        assert result[0]["remarks"] == ""
    
    def test_merged_cell_handling(self):
        """
        TC-2.4: 合并单元格处理测试。
        
        验证合并单元格取首值。
        """
        # Arrange
        from src.services.parsers.excel_parser import ExcelParser
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A1:B1")
        ws["A1"] = "合并值"
        ws["A2"] = "正常值"
        
        # Act
        parser = ExcelParser()
        result = parser.parse_workbook(wb)
        
        # Assert
        assert result[0]["A"] == "合并值"


# =============================================================================
# TC-6: CSV 导出功能测试 (ATB-3)
# =============================================================================

class TestCsvExport:
    """
    CSV 导出功能测试类。
    验证基础导出、筛选导出、大数据量导出等场景。
    """
    
    def test_export_all_assets_to_csv(self, mock_asset_service):
        """
        TC-3.1: 基础 CSV 导出测试。
        
        验证能够正确导出所有资产数据。
        """
        # Arrange
        from src.services.export_service import CsvExportService
        
        mock_assets = [
            {"asset_name": "资产A", "asset_type": "IT_HARDWARE", "status": "ACTIVE"},
            {"asset_name": "资产B", "asset_type": "FURNITURE", "status": "ACTIVE"},
        ]
        mock_asset_service.export_assets.return_value = mock_assets
        
        # Act
        service = CsvExportService(asset_service=mock_asset_service)
        result = service.export()
        
        # Assert
        assert "asset_name,asset_type,status" in result
        assert "资产A,IT_HARDWARE,ACTIVE" in result
    
    def test_export_filtered_by_type(self, mock_asset_service):
        """
        TC-3.2: 筛选条件导出测试。
        
        验证能够按资产类型筛选导出。
        """
        # Arrange
        from src.services.export_service import CsvExportService
        
        filtered_assets = [
            {"asset_name": "设备A", "asset_type": "EQUIPMENT", "status": "ACTIVE"},
        ]
        mock_asset_service.export_assets.return_value = filtered_assets
        
        # Act
        service = CsvExportService(asset_service=mock_asset_service)
        result = service.export(filters={"asset_type": "EQUIPMENT"})
        
        # Assert
        assert "EQUIPMENT" in result
        assert "FURNITURE" not in result
    
    def test_export_10000_records(self, mock_asset_service):
        """
        TC-3.3: 大数据量导出测试。
        
        验证 10000 条记录的导出性能和内存使用。
        """
        # Arrange
        from src.services.export_service import CsvExportService
        import tracemalloc
        
        large_dataset = [
            {"asset_name": f"资产-{i}", "asset_type": "IT_HARDWARE", "status": "ACTIVE"}
            for i in range(10000)
        ]
        mock_asset_service.export_assets.return_value = large_dataset
        
        # Act
        tracemalloc.start()
        service = CsvExportService(asset_service=mock_asset_service)
        start_time = datetime.now()
        result = service.export()
        elapsed = (datetime.now() - start_time).total_seconds()
        current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        
        # Assert
        assert elapsed < 10.0  # 10秒内完成
        assert peak < 512 * 1024 * 1024  # 峰值内存 < 512MB
    
    def test_paginated_export(self, mock_asset_service):
        """
        TC-3.4: 分页导出测试。
        
        验证导出支持分页参数。
        """
        # Arrange
        from src.services.export_service import CsvExportService
        
        mock_asset_service.export_assets.return_value = []
        
        # Act
        service = CsvExportService(asset_service=mock_asset_service)
        result = service.export(page=1, page_size=100)
        
        # Assert
        mock_asset_service.export_assets.assert_called_once()


# =============================================================================
# TC-7: Excel 导出功能测试 (ATB-4)
# =============================================================================

class TestExcelExport:
    """
    Excel 导出功能测试类。
    验证 Excel 格式生成、列宽自适应、中文表头等场景。
    """
    
    def test_export_to_xlsx(self, mock_asset_service):
        """
        TC-4.1: Excel 格式导出测试。
        
        验证能够生成有效的 .xlsx 文件。
        """
        # Arrange
        from src.services.export_service import ExcelExportService
        from openpyxl import load_workbook
        
        mock_assets = [
            {"asset_name": "资产A", "asset_type": "IT_HARDWARE"},
        ]
        mock_asset_service.export_assets.return_value = mock_assets
        
        # Act
        service = ExcelExportService(asset_service=mock_asset_service)
        file_stream = service.export()
        
        # Assert
        wb = load_workbook(file_stream)
        ws = wb.active
        assert ws["A1"].value == "asset_name"
        assert ws["A2"].value == "资产A"
    
    def test_auto_adjust_column_width(self, mock_asset_service):
        """
        TC-4.2: 列宽自适应测试。
        
        验证列宽能够根据内容自动调整。
        """
        # Arrange
        from src.services.export_service import ExcelExportService
        
        mock_assets = [
            {"asset_name": "这是一个很长很长的资产名称需要自动调整宽度"},
        ]
        mock_asset_service.export_assets.return_value = mock_assets
        
        # Act
        service = ExcelExportService(asset_service=mock_asset_service)
        file_stream = service.export()
        
        # Assert
        from openpyxl import load_workbook
        wb = load_workbook(file_stream)
        ws = wb.active
        col_width = ws.column_dimensions["A"].width
        assert col_width > 10  # 列宽应大于默认值
    
    def test_chinese_column_headers(self, mock_asset_service):
        """
        TC-4.3: 中文表头支持测试。
        
        验证导出文件表头显示中文。
        """
        # Arrange
        from src.services.export_service import ExcelExportService
        
        mock_assets = [{"asset_name": "测试"}]
        mock_asset_service.export_assets.return_value = mock_assets
        
        # Act
        service = ExcelExportService(asset_service=mock_asset_service)
        file_stream = service.export(language="zh-CN")
        
        # Assert
        from openpyxl import load_workbook
        wb = load_workbook(file_stream)
        ws = wb.active
        assert "资产名称" in str(ws["A1"].value)


# =============================================================================
# TC-8: 前端 UI 集成测试 (ATB-5)
# =============================================================================

class TestImportUI:
    """
    导入功能前端 UI 测试类。
    使用 Playwright 进行端到端测试。
    """
    
    @pytest.fixture
    def browser_context(self, browser):
        """
        配置浏览器上下文。
        
        Args:
            browser: Playwright 浏览器实例
        
        Returns:
            BrowserContext: 配置好的浏览器上下文
        """
        context = browser.new_context()
        yield context
        context.close()
    
    def test_upload_csv_file(self, page):
        """
        TC-5.1: 文件上传交互测试。
        
        验证文件上传后触发预览并显示前10行。
        """
        # Arrange
        page.goto("/asset-bulk-import")
        
        # Act
        with page.expect_file_chooser() as fc_info:
            page.get_by_test_id("file-upload-input").click()
        file_chooser = fc_info.value
        file_chooser.set_files("tests/fixtures/sample_assets.csv")
        
        # Assert
        preview_table = page.get_by_test_id("data-preview-table")
        assert preview_table.is_visible()
        rows = preview_table.locator("tr")
        assert rows.count() <= 11  # header + 最多10行预览
    
    def test_upload_progress_indicator(self, page):
        """
        TC-5.2: 上传进度显示测试。
        
        验证上传过程中显示进度条。
        """
        # Arrange
        page.goto("/asset-bulk-import")
        
        # Act
        page.get_by_test_id("file-upload-input").set_input_files("tests/fixtures/large_import.csv")
        
        # Assert
        progress_bar = page.get_by_test_id("upload-progress")
        assert progress_bar.is_visible()
        # 完成后应显示结果
        page.wait_for_selector('[data-testid="import-result"]', timeout=60000)
    
    def test_validation_errors_display(self, page):
        """
        TC-5.3: 错误提示 UI 测试。
        
        验证校验失败时高亮错误行并可下载错误报告。
        """
        # Arrange
        page.goto("/asset-bulk-import")
        
        # Act
        page.get_by_test_id("file-upload-input").set_input_files("tests/fixtures/invalid_assets.csv")
        
        # Assert
        error_rows = page.locator('[data-testid="error-row"]')
        assert error_rows.count() > 0
        
        download_btn = page.get_by_test_id("download-error-report")
        assert download_btn.is_enabled()
    
    def test_export_button_click(self, page):
        """
        TC-5.4: 导出按钮交互测试。
        
        验证点击导出触发下载，文件名含时间戳。
        """
        # Arrange
        page.goto("/asset-list")
        
        # Act
        page.get_by_test_id("export-btn").click()
        
        # Assert
        with page.expect_download() as download_info:
            page.get_by_test_id("confirm-export").click()
        
        download = download_info.value
        assert "assets_export_" in download.suggested_filename
        assert datetime.now().strftime("%Y%m%d") in download.suggested_filename
    
    def test_export_format_selector(self, page):
        """
        TC-5.5: 导出格式选择测试。
        
        验证可切换 CSV/Excel 格式并刷新预览。
        """
        # Arrange
        page.goto("/asset-list")
        
        # Act
        csv_option = page.get_by_test_id("format-csv")
        excel_option = page.get_by_test_id("format-excel")
        
        csv_option.click()
        assert csv_option.is_checked()
        
        excel_option.click()
        assert excel_option.is_checked()


# =============================================================================
# TC-9: 集成测试 - API 端点测试
# =============================================================================

class TestAssetBulkApi:
    """
    资产批量操作 API 端点集成测试类。
    验证上传、查询、导出等 API 端点。
    """
    
    def test_upload_endpoint_returns_task_id(self, client):
        """
        验证上传端点返回任务ID。
        
        测试导入文件上传 API 返回正确的任务标识。
        """
        # Arrange
        files = {"file": ("test.csv", open("tests/fixtures/sample.csv", "rb"), "text/csv")}
        
        # Act
        response = client.post("/api/v1/assets/import/upload", files=files)
        
        # Assert
        assert response.status_code == 200
        data = response.json()
        assert "task_id" in data
        assert "status" in data
    
    def test_list_import_tasks(self, client):
        """
        验证导入任务列表 API。
        
        测试获取当前用户导入任务列表。
        """
        # Act
        response = client.get("/api/v1/assets/import/tasks")
        
        # Assert
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
    
    def test_get_task_status(self, client):
        """
        验证任务状态查询 API。
        
        测试通过任务ID查询导入状态。
        """
        # Act
        response = client.get("/api/v1/assets/import/tasks/task-001")
        
        # Assert
        assert response.status_code == 200
        data = response.json()
        assert data["task_id"] == "task-001"
    
    def test_export_endpoint(self, client):
        """
        验证资产导出 API。
        
        测试 CSV 和 Excel 格式导出。
        """
        # Act
        csv_response = client.get("/api/v1/assets/export?format=csv")
        excel_response = client.get("/api/v1/assets/export?format=xlsx")
        
        # Assert
        assert csv_response.status_code == 200
        assert csv_response.headers["content-type"] == "text/csv"
        assert excel_response.status_code == 200
        assert "spreadsheet" in excel_response.headers["content-type"]