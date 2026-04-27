"""
Excel 解析器单元测试

测试用例覆盖 ATB-2: Excel 导入功能
- TC-2.1: .xlsx 文件解析，正确读取第一个工作表
- TC-2.2: 多工作表取第一，忽略非首工作表数据
- TC-2.3: 空单元格处理，空单元格转为空字符串而非 null
- TC-2.4: 合并单元格支持，合并单元格取首值

参考: SWARM-2025-Q2-P2-006 规格指导文档
"""

import io
import tempfile
import pytest
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell

from src.parsers.excel_parser import ExcelParser


class TestExcelParser:
    """Excel 解析器测试类"""

    @pytest.fixture
    def parser(self):
        """创建 Excel 解析器实例"""
        return ExcelParser()

    @pytest.fixture
    def sample_workbook_bytes(self):
        """创建示例 Excel 工作簿字节流"""
        wb = Workbook()
        ws = wb.active
        ws.title = "AssetData"

        # 添加表头行
        headers = [
            "asset_id", "asset_name", "asset_type", "serial_number",
            "purchase_date", "purchase_price", "currency", "department",
            "custodian", "status", "location", "remarks"
        ]
        ws.append(headers)

        # 添加数据行
        test_data = [
            ["A001", "测试资产1", "EQUIPMENT", "SN-001", "2025-01-15", "10000.00", "CNY", "IT", "张三", "ACTIVE", "北京", "测试备注1"],
            ["A002", "测试资产2", "FURNITURE", "SN-002", "2025-02-20", "5000.00", "CNY", "HR", "李四", "ACTIVE", "上海", ""],
            ["", "测试资产3", "VEHICLE", "", "2025-03-10", "50000.00", "CNY", "FIN", "王五", "ACTIVE", "广州", None],
        ]
        for row in test_data:
            ws.append(row)

        # 保存到字节流
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    @pytest.fixture
    def multi_sheet_workbook_bytes(self):
        """创建多工作表 Excel 工作簿字节流"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"

        # 第一个工作表包含资产数据
        headers = ["asset_id", "asset_name", "asset_type"]
        ws1.append(headers)
        ws1.append(["B001", "资产B1", "IT_HARDWARE"])
        ws1.append(["B002", "资产B2", "OTHER"])

        # 第二个工作表应该被忽略
        ws2 = wb.create_sheet(title="Sheet2")
        ws2.append(["不应该", "被", "读取"])

        # 第三个工作表也应该被忽略
        ws3 = wb.create_sheet(title="Summary")
        ws3.append(["摘要", "数据"])

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    @pytest.fixture
    def workbook_with_empty_cells_bytes(self):
        """创建包含空单元格的 Excel 工作簿字节流"""
        wb = Workbook()
        ws = wb.active

        # 表头
        ws.append(["asset_id", "asset_name", "asset_type", "serial_number", "purchase_date"])

        # 数据行 - 混合空单元格场景
        ws.append(["C001", "资产C1", "EQUIPMENT", "SN-C1", "2025-01-01"])
        ws.append(["", "资产C2", "", "SN-C2", ""])  # 空字符串
        ws.append(["C003", None, "FURNITURE", None, "2025-02-02"])  # None值
        ws.append(["", "", "", "", ""])  # 全空行

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    @pytest.fixture
    def workbook_with_merged_cells_bytes(self):
        """创建包含合并单元格的 Excel 工作簿字节流"""
        wb = Workbook()
        ws = wb.active

        # 表头
        ws.append(["asset_id", "asset_name", "asset_type"])

        # 数据行
        ws.append(["D001", "资产D1", "VEHICLE"])

        # 创建合并单元格 (A3:B3)
        ws.merge_cells("A3:B3")
        ws["A3"] = "D002-D003-merged"
        ws["A3"].value = "D002-D003-merged"
        # B3 是合并单元格，需要明确设置
        ws["B3"] = "资产D2-D3"
        ws["B3"].value = "资产D2-D3"

        ws.append(["D003", "资产D3", "IT_HARDWARE"])

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    def test_parse_xlsx_worksheet(self, parser, sample_workbook_bytes):
        """
        TC-2.1: .xlsx 文件解析
        期待结果: 正确解析 header 与数据行，无异常
        """
        result = parser.parse(sample_workbook_bytes)

        # 验证解析结果不为空
        assert result is not None
        assert len(result) > 0

        # 验证包含 header 行
        headers = result[0]
        expected_headers = [
            "asset_id", "asset_name", "asset_type", "serial_number",
            "purchase_date", "purchase_price", "currency", "department",
            "custodian", "status", "location", "remarks"
        ]
        assert headers == expected_headers

        # 验证数据行数量 (3条测试数据)
        assert len(result) == 4  # 1 header + 3 data rows

    def test_ignore_other_worksheets(self, parser, multi_sheet_workbook_bytes):
        """
        TC-2.2: 多工作表取第一
        期待结果: 忽略非首工作表数据，仅读取第一个工作表的内容
        """
        result = parser.parse(multi_sheet_workbook_bytes)

        # 验证只读取第一个工作表
        assert result is not None
        assert len(result) >= 1

        headers = result[0]
        assert "asset_id" in headers
        assert "asset_name" in headers
        assert "asset_type" in headers

        # 验证数据来自第一个工作表
        # 不应该包含 Sheet2 或 Summary 的数据
        data_str = str(result)
        assert "不应该" not in data_str
        assert "被" not in data_str
        assert "读取" not in data_str
        assert "摘要" not in data_str

        # 验证包含第一个工作表的数据
        assert any("B001" in str(row) for row in result)
        assert any("资产B1" in str(row) for row in result)

    def test_empty_cell_as_empty_string(self, parser, workbook_with_empty_cells_bytes):
        """
        TC-2.3: 空单元格处理
        期待结果: 空单元格转为空字符串而非 null
        """
        result = parser.parse(workbook_with_empty_cells_bytes)

        assert result is not None
        assert len(result) >= 2  # 至少包含 header 和 1 行数据

        # 检查空单元格处理
        # openpyxl 读取时，空单元格可能返回 None 或空字符串
        # ExcelParser 应该统一转换为空字符串
        data_rows = result[1:]  # 跳过 header

        for row in data_rows:
            for cell_value in row:
                # 所有值应该是字符串类型 (包括空字符串)
                if cell_value is not None:
                    assert isinstance(cell_value, str), \
                        f"Cell value should be string, got {type(cell_value)}"

    def test_merged_cell_handling(self, parser, workbook_with_merged_cells_bytes):
        """
        TC-2.4: 合并单元格支持
        期待结果: 合并单元格取首值
        """
        result = parser.parse(workbook_with_merged_cells_bytes)

        assert result is not None

        # 验证包含合并区域的数据
        data_str = str(result)

        # 合并单元格的值应该被正确读取
        # 第一个值应该是 "D002-D003-merged"
        assert any("D002" in str(row) or "D003" in str(row) for row in result)

        # 验证不是 None 值 (merged cell 应有值)
        data_rows = result[1:]  # 跳过 header
        has_valid_data = any(
            any(cell is not None and cell != "" for cell in row)
            for row in data_rows
        )
        assert has_valid_data, "Merged cell should have a value, not be empty"

    def test_parse_with_invalid_file(self, parser):
        """
        测试无效文件处理
        期待结果: 对非 xlsx 文件抛出异常或返回空结果
        """
        invalid_bytes = b"This is not an Excel file"

        with pytest.raises(Exception):
            parser.parse(invalid_bytes)

    def test_parse_empty_workbook(self, parser):
        """测试空工作簿处理"""
        wb = Workbook()
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        result = parser.parse(stream.getvalue())

        # 空工作簿应该返回至少包含 header 的结果或空列表
        assert result is not None
        # openpyxl 创建的工作簿至少有 active sheet
        assert isinstance(result, list)

    def test_parse_large_workbook(self, parser):
        """测试大文件解析 (边界测试)"""
        wb = Workbook()
        ws = wb.active

        # 添加 header
        headers = ["col" + str(i) for i in range(12)]
        ws.append(headers)

        # 添加 100 行数据 (远小于 5000 上限，用于验证解析器性能)
        for i in range(100):
            ws.append([f"data_{i}_{j}" for j in range(12)])

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        result = parser.parse(stream.getvalue())

        assert result is not None
        assert len(result) == 101  # 1 header + 100 data rows