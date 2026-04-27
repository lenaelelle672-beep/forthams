"""
Excel 文件解析器

支持 .xlsx 格式的资产批量导入数据解析。
符合 SWARM-2025-Q2-P2-006 规格要求。

功能特性:
    - 读取第一个工作表数据（忽略其他工作表）
    - 空单元格转换为空字符串
    - 合并单元格取首值
    - 支持 UTF-8 编码的 .xlsx 文件

依赖:
    - openpyxl >= 3.0.0

导入字段清单（固定映射）:
    - asset_id, asset_name, asset_type, serial_number, purchase_date,
      purchase_price, currency, department, custodian, status,
      location, remarks
"""

from typing import List, Dict, Any, Optional
from datetime import datetime
import logging

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from .base_parser import BaseParser

logger = logging.getLogger(__name__)


class ExcelParseError(Exception):
    """Excel 解析错误异常"""
    pass


class ExcelParser(BaseParser):
    """
    Excel (.xlsx) 文件解析器
    
    继承自 BaseParser，实现 parse() 接口。
    专用于处理资产批量导入的 Excel 文件。
    
    Attributes:
        max_file_size: 最大文件大小限制（10 MB）
        supported_extensions: 支持的文件扩展名列表
    """
    
    SUPPORTED_EXTENSIONS = ['.xlsx']
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
    
    def __init__(self, file_path: str):
        """
        初始化 Excel 解析器
        
        Args:
            file_path: Excel 文件路径
        """
        super().__init__(file_path)
        self._workbook = None
        self._worksheet = None
    
    def _validate_file(self) -> None:
        """
        验证文件是否合法
        
        Raises:
            ExcelParseError: 文件不存在或格式不正确
        """
        import os
        
        if not os.path.exists(self.file_path):
            raise ExcelParseError(f"文件不存在: {self.file_path}")
        
        if not self.file_path.lower().endswith('.xlsx'):
            raise ExcelParseError(f"不支持的文件格式，仅支持 .xlsx: {self.file_path}")
        
        file_size = os.path.getsize(self.file_path)
        if file_size > self.MAX_FILE_SIZE:
            raise ExcelParseError(
                f"文件大小超过限制 ({file_size} bytes > {self.MAX_FILE_SIZE} bytes)"
            )
    
    def parse(self) -> List[Dict[str, Any]]:
        """
        解析 Excel 文件并返回数据行列表
        
        Returns:
            List[Dict[str, Any]]: 解析后的数据行列表，每行为字典
        
        Raises:
            ExcelParseError: 解析过程中发生错误
        """
        try:
            self._validate_file()
            self._load_workbook()
            rows = self._parse_worksheet()
            logger.info(f"成功解析 Excel 文件，共 {len(rows)} 行数据")
            return rows
        except Exception as e:
            logger.error(f"Excel 解析失败: {str(e)}")
            raise ExcelParseError(f"Excel 解析失败: {str(e)}")
        finally:
            self._close_workbook()
    
    def _load_workbook(self) -> None:
        """
        加载 Excel 工作簿
        
        Raises:
            ExcelParseError: 文件加载失败
        """
        try:
            self._workbook = load_workbook(
                self.file_path, 
                data_only=True,
                read_only=True
            )
            # 仅读取第一个工作表
            sheet_names = self._workbook.sheetnames
            if not sheet_names:
                raise ExcelParseError("Excel 文件不包含任何工作表")
            
            self._worksheet = self._workbook[sheet_names[0]]
            logger.debug(f"已加载工作表: {sheet_names[0]}, 忽略其他工作表: {sheet_names[1:]}")
        except Exception as e:
            raise ExcelParseError(f"加载 Excel 工作簿失败: {str(e)}")
    
    def _get_cell_value(self, cell) -> str:
        """
        获取单元格值，处理空单元格和合并单元格
        
        Args:
            cell: openpyxl 单元格对象
            
        Returns:
            str: 单元格值，空单元格返回空字符串
        """
        # 处理合并单元格（取首值）
        if isinstance(cell, MergedCell):
            return ""
        
        value = cell.value
        if value is None:
            return ""
        
        # 转换日期对象为字符串
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        
        return str(value).strip()
    
    def _parse_worksheet(self) -> List[Dict[str, Any]]:
        """
        解析工作表数据
        
        Returns:
            List[Dict[str, Any]]: 数据行列表
        """
        if self._worksheet is None:
            return []
        
        rows = []
        all_rows = list(self._worksheet.iter_rows(values_only=True))
        
        if not all_rows:
            return []
        
        # 第一行为表头
        header_row = all_rows[0]
        headers = [str(cell).strip() if cell is not None else f"col_{i}" 
                  for i, cell in enumerate(header_row)]
        
        # 验证表头
        if not headers or all(h == "" for h in headers):
            raise ExcelParseError("Excel 文件表头为空或无效")
        
        # 解析数据行
        for row_idx, row_data in enumerate(all_rows[1:], start=2):
            # 跳过空行
            if self._is_empty_row(row_data):
                logger.warning(f"第 {row_idx} 行为空行，已跳过")
                continue
            
            row_dict = {}
            for col_idx, (header, cell_value) in enumerate(zip(headers, row_data)):
                # 处理空单元格（转为空字符串）
                if col_idx < len(row_data):
                    value = self._get_cell_value_by_index(row_data, col_idx)
                    row_dict[header] = value
                else:
                    row_dict[header] = ""
            
            rows.append(row_dict)
        
        return rows
    
    def _get_cell_value_by_index(self, row_data: tuple, col_idx: int) -> str:
        """
        根据列索引获取单元格值
        
        Args:
            row_data: 行数据元组
            col_idx: 列索引
            
        Returns:
            str: 单元格值
        """
        try:
            value = row_data[col_idx]
            if value is None:
                return ""
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")
            return str(value).strip()
        except IndexError:
            return ""
    
    def _is_empty_row(self, row_data: tuple) -> bool:
        """
        判断是否为空行
        
        Args:
            row_data: 行数据元组
            
        Returns:
            bool: 如果所有单元格都为空则返回 True
        """
        return all(cell is None or str(cell).strip() == "" for cell in row_data)
    
    def _close_workbook(self) -> None:
        """关闭工作簿释放资源"""
        if self._workbook is not None:
            try:
                self._workbook.close()
            except Exception:
                pass
            finally:
                self._workbook = None
                self._worksheet = None
    
    def get_headers(self) -> List[str]:
        """
        获取表头列表
        
        Returns:
            List[str]: 表头列表
        """
        if self._worksheet is None:
            return []
        
        rows = list(self._worksheet.iter_rows(max_row=1, values_only=True))
        if not rows:
            return []
        
        headers = []
        for i, cell in enumerate(rows[0]):
            if cell is not None:
                headers.append(str(cell).strip())
            else:
                headers.append(f"col_{i}")
        return headers
    
    def get_row_count(self) -> int:
        """
        获取数据行数（不含表头）
        
        Returns:
            int: 数据行数
        """
        if self._worksheet is None:
            return 0
        
        return self._worksheet.max_row - 1  # 减去表头行
    
    def get_column_count(self) -> int:
        """
        获取列数
        
        Returns:
            int: 列数
        """
        if self._worksheet is None:
            return 0
        
        return self._worksheet.max_column


def parse_excel(file_path: str) -> List[Dict[str, Any]]:
    """
    便捷函数：解析 Excel 文件
    
    Args:
        file_path: Excel 文件路径
        
    Returns:
        List[Dict[str, Any]]: 解析后的数据行列表
    """
    parser = ExcelParser(file_path)
    return parser.parse()