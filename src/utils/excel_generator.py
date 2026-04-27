"""
资产批量导入导出工具模块 - Excel 生成器

本模块提供资产数据的 Excel/CSV 文件生成功能，支持：
- 导入模板生成（xlsx/xlsx/csv）
- 资产数据批量导出
- 导入错误报告生成

功能标识: SWARM-2025-Q2-P2-006 (Iteration 2)

支持的字段:
    - asset_id: 资产编号（唯一标识）
    - asset_name: 资产名称
    - asset_type: 资产类型（EQUIPMENT/INSTRUMENT/VEHICLE/OTHER）
    - purchase_date: 采购日期（YYYY-MM-DD）
    - purchase_amount: 采购金额（≥0，精度2位）
    - department: 所属部门编码
    - status: 资产状态（ACTIVE/MAINTENANCE/SCRAPPED）
    - location: 存放地点（可选）
    - description: 资产描述（可选）

约束限制:
    - 单次导出行数上限: 500,000 行
    - 文件大小建议不超过 50MB
    - 导出链接有效期: 24 小时

典型用法:
    >>> from src.utils.excel_generator import ExcelGenerator
    >>> generator = ExcelGenerator()
    >>> # 生成导入模板
    >>> template_path = generator.generate_import_template("xlsx")
    >>> # 导出资产数据
    >>> export_path = generator.export_assets(assets, "xlsx")
    >>> # 生成错误报告
    >>> error_report = generator.generate_error_report(errors, "xlsx")
"""

import io
import csv
import logging
from datetime import datetime, date
from decimal import Decimal
from typing import List, Dict, Any, Optional, Union, Tuple
from enum import Enum

# 尝试导入 openpyxl，如果不可用则使用 xlsxwriter 作为备选
try:
    from openpyxl import Workbook, styles
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_ENGINE = "openpyxl"
except ImportError:
    try:
        import xlsxwriter
        EXCEL_ENGINE = "xlsxwriter"
    except ImportError:
        EXCEL_ENGINE = None
        logging.warning("Neither openpyxl nor xlsxwriter is available. Excel generation will be limited.")

from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


class ExportFormat(Enum):
    """支持的导出文件格式枚举"""
    XLSX = "xlsx"
    XLS = "xls"
    CSV = "csv"


class TemplateType(Enum):
    """导入模板类型枚举"""
    ASSET_IMPORT = "asset_import"
    ERROR_REPORT = "error_report"


@dataclass
class AssetField:
    """资产字段定义数据类
    
    Attributes:
        name: 字段英文名称
        display_name: 字段中文显示名称
        required: 是否必填
        max_length: 最大长度限制
        validation_type: 校验类型（enum/date/number/string）
        allowed_values: 枚举值列表（当 validation_type 为 enum 时）
        description: 字段说明
    """
    name: str
    display_name: str
    required: bool = True
    max_length: Optional[int] = None
    validation_type: str = "string"
    allowed_values: Optional[List[str]] = None
    description: str = ""


@dataclass
class ImportError:
    """导入错误信息数据类
    
    Attributes:
        row: 错误所在行号
        field: 错误字段名
        message: 错误描述信息
        original_value: 原始值
    """
    row: int
    field: str
    message: str
    original_value: Optional[str] = None


@dataclass
class ExportOptions:
    """导出选项配置数据类
    
    Attributes:
        format: 导出文件格式
        include_header: 是否包含表头行
        sheet_name: 工作表名称
        date_format: 日期字段格式
        decimal_places: 金额字段小数位数
        encoding: CSV 文件编码格式
    """
    format: ExportFormat = ExportFormat.XLSX
    include_header: bool = True
    sheet_name: str = "资产数据"
    date_format: str = "%Y-%m-%d"
    decimal_places: int = 2
    encoding: str = "utf-8-sig"


class ExcelGenerator:
    """
    Excel 文件生成器类
    
    提供资产数据的 Excel/CSV 文件生成功能，包括：
    - 导入模板生成
    - 数据批量导出
    - 错误报告生成
    
    典型用法:
        >>> generator = ExcelGenerator()
        >>> template = generator.generate_import_template(ExportFormat.XLSX)
        >>> # 或指定文件保存路径
        >>> generator.generate_import_template(ExportFormat.XLSX, output_path="/tmp/template.xlsx")
    
    Attributes:
        ASSET_FIELDS: 资产字段定义列表
    """
    
    # 资产字段定义 - 与后端数据模型保持一致
    ASSET_FIELDS: List[AssetField] = [
        AssetField(
            name="asset_id",
            display_name="资产编号",
            required=True,
            max_length=64,
            validation_type="string",
            description="唯一标识资产，字母数字下划线"
        ),
        AssetField(
            name="asset_name",
            display_name="资产名称",
            required=True,
            max_length=128,
            validation_type="string",
            description="资产的中文或英文名称"
        ),
        AssetField(
            name="asset_type",
            display_name="资产类型",
            required=True,
            validation_type="enum",
            allowed_values=["EQUIPMENT", "INSTRUMENT", "VEHICLE", "OTHER"],
            description="EQUIPMENT-设备 INSTRUMENT-仪器 VEHICLE-车辆 OTHER-其他"
        ),
        AssetField(
            name="purchase_date",
            display_name="采购日期",
            required=True,
            validation_type="date",
            description="采购日期，格式 YYYY-MM-DD"
        ),
        AssetField(
            name="purchase_amount",
            display_name="采购金额",
            required=True,
            validation_type="number",
            description="采购金额，必须 ≥0，精度2位小数"
        ),
        AssetField(
            name="department",
            display_name="所属部门",
            required=True,
            max_length=64,
            validation_type="string",
            description="所属部门编码，需匹配已存在部门"
        ),
        AssetField(
            name="status",
            display_name="资产状态",
            required=True,
            validation_type="enum",
            allowed_values=["ACTIVE", "MAINTENANCE", "SCRAPPED"],
            description="ACTIVE-在用 MAINTENANCE-维护中 SCRAPPED-已报废"
        ),
        AssetField(
            name="location",
            display_name="存放地点",
            required=False,
            max_length=128,
            validation_type="string",
            description="资产的物理存放位置"
        ),
        AssetField(
            name="description",
            display_name="资产描述",
            required=False,
            max_length=512,
            validation_type="string",
            description="资产的详细描述信息"
        ),
    ]
    
    # 表头样式定义
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    REQUIRED_MARKER_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    # 普通单元格样式
    CELL_FONT = Font(name="微软雅黑", size=10)
    CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    
    # 边框样式
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def __init__(self) -> None:
        """
        初始化 Excel 生成器
        
        检查可用的 Excel 处理库并初始化配置
        """
        self._engine = EXCEL_ENGINE
        logger.info(f"ExcelGenerator initialized with engine: {self._engine}")
        
    def _validate_format(self, format_type: Union[str, ExportFormat]) -> ExportFormat:
        """
        验证并标准化导出格式
        
        Args:
            format_type: 格式标识（字符串或枚举）
            
        Returns:
            ExportFormat: 标准化的格式枚举
            
        Raises:
            ValueError: 不支持的格式类型
        """
        if isinstance(format_type, str):
            format_type = format_type.lower().strip()
            if format_type == "xlsx":
                return ExportFormat.XLSX
            elif format_type == "xls":
                return ExportFormat.XLS
            elif format_type == "csv":
                return ExportFormat.CSV
            else:
                raise ValueError(f"Unsupported format: {format_type}")
        elif isinstance(format_type, ExportFormat):
            return format_type
        else:
            raise ValueError(f"Format must be str or ExportFormat, got {type(format_type)}")

    def _create_workbook(self) -> Any:
        """
        创建工作簿对象
        
        Returns:
            Workbook: 工作簿实例（openpyxl 或 xlsxwriter）
        """
        if self._engine == "openpyxl":
            return Workbook()
        elif self._engine == "xlsxwriter":
            return xlsxwriter.Workbook()
        else:
            raise RuntimeError("No Excel engine available. Please install openpyxl or xlsxwriter.")

    def generate_import_template(
        self,
        format_type: Union[str, ExportFormat] = ExportFormat.XLSX,
        output_path: Optional[str] = None,
        include_sample: bool = True,
        sheet_name: str = "资产导入模板"
    ) -> Union[str, bytes]:
        """
        生成资产批量导入模板文件
        
        生成符合规范的 Excel/CSV 模板文件，包含：
        - 表头行（字段名称 + 必填标记）
        - 示例数据行（可选）
        - 字段说明注释
        
        Args:
            format_type: 导出格式，支持 'xlsx', 'xls', 'csv'
            output_path: 文件保存路径，为 None 时返回字节数据
            include_sample: 是否包含示例数据行
            sheet_name: 工作表名称（仅 Excel 格式有效）
            
        Returns:
            str: 当 output_path 指定时返回文件路径
            bytes: 当 output_path 为 None 时返回文件内容
            
        Raises:
            ValueError: 不支持的格式类型
            RuntimeError: Excel 引擎不可用
            
        典型用法:
            >>> generator = ExcelGenerator()
            >>> # 生成 xlsx 模板到指定路径
            >>> path = generator.generate_import_template("xlsx", "/tmp/template.xlsx")
            >>> # 生成 csv 模板返回字节数据
            >>> data = generator.generate_import_template("csv")
        """
        format_enum = self._validate_format(format_type)
        
        logger.info(f"Generating import template: format={format_enum.value}, output={output_path}")
        
        if format_enum == ExportFormat.CSV:
            return self._generate_csv_template(output_path, include_sample)
        else:
            return self._generate_excel_template(format_enum, output_path, include_sample, sheet_name)

    def _generate_excel_template(
        self,
        format_enum: ExportFormat,
        output_path: Optional[str],
        include_sample: bool,
        sheet_name: str
    ) -> Union[str, bytes]:
        """
        生成 Excel 格式的导入模板
        
        Args:
            format_enum: 格式枚举
            output_path: 输出路径
            include_sample: 是否包含示例
            sheet_name: 工作表名称
            
        Returns:
            文件路径或字节数据
        """
        if self._engine == "openpyxl":
            return self._generate_with_openpyxl(format_enum, output_path, include_sample, sheet_name)
        elif self._engine == "xlsxwriter":
            return self._generate_with_xlsxwriter(format_enum, output_path, include_sample, sheet_name)
        else:
            raise RuntimeError("Excel engine not available")

    def _generate_with_openpyxl(
        self,
        format_enum: ExportFormat,
        output_path: Optional[str],
        include_sample: bool,
        sheet_name: str
    ) -> Union[str, bytes]:
        """
        使用 openpyxl 生成 Excel 模板
        
        Args:
            format_enum: 格式枚举
            output_path: 输出路径
            include_sample: 是否包含示例
            sheet_name: 工作表名称
            
        Returns:
            文件路径或字节数据
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 设置默认列宽
        column_widths = {
            "A": 15,  # asset_id
            "B": 20,  # asset_name
            "C": 15,  # asset_type
            "D": 15,  # purchase_date
            "E": 15,  # purchase_amount
            "F": 15,  # department
            "G": 15,  # status
            "H": 20,  # location
            "I": 30,  # description
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 写入表头
        for col_idx, field in enumerate(self.ASSET_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field.display_name
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL if field.required else PatternFill(fill_type=None)
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
            
            # 必填字段标记
            if field.required:
                cell.comment = None  # 可扩展添加批注
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # 写入示例数据
        if include_sample:
            sample_data = self._get_sample_data()
            for row_idx, row_data in enumerate(sample_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = self.CELL_FONT
                    cell.alignment = self.CELL_ALIGNMENT
                    cell.border = self.THIN_BORDER
                    # 示例数据使用灰色字体标记
                    cell.font = Font(name="微软雅黑", size=10, color="808080")
        
        # 添加字段说明工作表
        self._add_field_description_sheet(wb)
        
        # 保存文件
        if output_path:
            wb.save(output_path)
            logger.info(f"Template saved to: {output_path}")
            return output_path
        else:
            # 返回字节数据
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

    def _generate_with_xlsxwriter(
        self,
        format_enum: ExportFormat,
        output_path: Optional[str],
        include_sample: bool,
        sheet_name: str
    ) -> Union[str, bytes]:
        """
        使用 xlsxwriter 生成 Excel 模板
        
        Args:
            format_enum: 格式枚举
            output_path: 输出路径
            include_sample: 是否包含示例
            sheet_name: 工作表名称
            
        Returns:
            文件路径或字节数据
        """
        if not output_path:
            raise ValueError("xlsxwriter requires output_path to be specified")
        
        wb = xlsxwriter.Workbook(output_path)
        ws = wb.add_worksheet(sheet_name)
        
        # 定义格式
        header_format = wb.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '4472C4',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        required_format = wb.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': 'C00000',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        cell_format = wb.add_format({
            'align': 'left',
            'valign': 'vcenter',
            'border': 1
        })
        
        sample_format = wb.add_format({
            'align': 'left',
            'valign': 'vcenter',
            'border': 1,
            'font_color': '808080'
        })
        
        # 写入表头
        for col_idx, field in enumerate(self.ASSET_FIELDS):
            fmt = required_format if field.required else header_format
            ws.write(0, col_idx, field.display_name, fmt)
            ws.set_column(col_idx, col_idx, 20)
        
        # 写入示例数据
        if include_sample:
            sample_data = self._get_sample_data()
            for row_idx, row_data in enumerate(sample_data, start=2):
                for col_idx, value in enumerate(row_data):
                    ws.write(row_idx - 1, col_idx, value, sample_format)
        
        wb.close()
        logger.info(f"Template saved to: {output_path}")
        return output_path

    def _generate_csv_template(
        self,
        output_path: Optional[str],
        include_sample: bool
    ) -> Union[str, bytes]:
        """
        生成 CSV 格式的导入模板
        
        Args:
            output_path: 输出路径，为 None 时返回字符串
            include_sample: 是否包含示例
            
        Returns:
            文件路径或 CSV 字符串
        """
        output = io.StringIO()
        
        # 写入 BOM 以支持 Excel 正确识别 UTF-8
        if output_path:
            # 使用二进制写入模式添加 BOM
            with open(output_path, 'wb') as f:
                f.write('\ufeff'.encode('utf-8-sig'))
            # 追加模式写入内容
            with open(output_path, 'a', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                self._write_csv_content(writer, include_sample)
            return output_path
        else:
            output.write('\ufeff')  # BOM for Excel
            writer = csv.writer(output)
            self._write_csv_content(writer, include_sample)
            return output.getvalue()

    def _write_csv_content(self, writer, include_sample: bool) -> None:
        """
        写入 CSV 内容
        
        Args:
            writer: CSV writer 对象
            include_sample: 是否包含示例数据
        """
        # 写入表头
        headers = [field.display_name for field in self.ASSET_FIELDS]
        writer.writerow(headers)
        
        # 写入示例数据
        if include_sample:
            sample_data = self._get_sample_data()
            for row_data in sample_data:
                writer.writerow(row_data)

    def _get_sample_data(self) -> List[List[Any]]:
        """
        获取示例数据行
        
        Returns:
            List[List[Any]]: 示例数据列表
        """
        today = datetime.now().strftime("%Y-%m-%d")
        return [
            [
                "AST-2025-00001",  # asset_id
                "办公笔记本电脑",    # asset_name
                "EQUIPMENT",       # asset_type
                today,             # purchase_date
                5999.99,           # purchase_amount
                "IT-DEPT-001",     # department
                "ACTIVE",          # status
                "3楼301室",        # location
                "ThinkPad T490, 16GB RAM, 512GB SSD",  # description
            ],
            [
                "AST-2025-00002",  # asset_id
                "激光打印机",       # asset_name
                "EQUIPMENT",       # asset_type
                today,             # purchase_date
                2999.00,           # purchase_amount
                "IT-DEPT-001",     # department
                "ACTIVE",          # status
                "3楼打印室",       # location
                "HP LaserJet Pro", # description
            ],
        ]

    def _add_field_description_sheet(self, wb) -> None:
        """
        添加字段说明工作表
        
        Args:
            wb: openpyxl Workbook 对象
        """
        ws = wb.create_sheet(title="字段说明")
        
        # 表头
        headers = ["字段名", "中文名", "必填", "数据类型", "约束", "说明"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # 写入字段说明
        for row_idx, field in enumerate(self.ASSET_FIELDS, start=2):
            ws.cell(row=row_idx, column=1).value = field.name
            ws.cell(row=row_idx, column=2).value = field.display_name
            ws.cell(row=row_idx, column=3).value = "是" if field.required else "否"
            ws.cell(row=row_idx, column=4).value = field.validation_type
            
            constraint = ""
            if field.allowed_values:
                constraint = ", ".join(field.allowed_values)
            elif field.max_length:
                constraint = f"最大 {field.max_length} 字符"
            ws.cell(row=row_idx, column=5).value = constraint
            ws.cell(row=row_idx, column=6).value = field.description

    def export_assets(
        self,
        assets: List[Dict[str, Any]],
        format_type: Union[str, ExportFormat] = ExportFormat.XLSX,
        output_path: Optional[str] = None,
        options: Optional[ExportOptions] = None
    ) -> Union[str, bytes]:
        """
        批量导出资产数据到文件
        
        将资产列表导出为 Excel 或 CSV 格式文件，支持：
        - 指定导出格式
        - 自定义列映射
        - 日期/金额格式化
        
        Args:
            assets: 资产数据列表，每项为字典
            format_type: 导出格式，支持 'xlsx', 'xls', 'csv'
            output_path: 文件保存路径，为 None 时返回字节数据
            options: 导出选项配置
            
        Returns:
            str: 当 output_path 指定时返回文件路径
            bytes: 当 output_path 为 None 时返回文件内容
            
        Raises:
            ValueError: 数据为空或格式不支持
            RuntimeError: Excel 引擎不可用
            
        典型用法:
            >>> generator = ExcelGenerator()
            >>> assets = [
            ...     {"asset_id": "AST-001", "asset_name": "设备A", ...},
            ...     {"asset_id": "AST-002", "asset_name": "设备B", ...},
            ... ]
            >>> path = generator.export_assets(assets, "xlsx", "/tmp/export.xlsx")
            >>> # 或指定选项
            >>> opts = ExportOptions(format=ExportFormat.CSV, date_format="%Y/%m/%d")
            >>> data = generator.export_assets(assets, "csv", options=opts)
        """
        if not assets:
            raise ValueError("Asset list cannot be empty")
        
        if len(assets) > 500000:
            logger.warning(f"Asset count {len(assets)} exceeds recommended limit of 500000")
        
        format_enum = self._validate_format(format_type)
        options = options or ExportOptions(format=format_enum)
        
        logger.info(f"Exporting {len(assets)} assets: format={format_enum.value}")
        
        if format_enum == ExportFormat.CSV:
            return self._export_csv(assets, output_path, options)
        else:
            return self._export_excel(assets, format_enum, output_path, options)

    def _export_excel(
        self,
        assets: List[Dict[str, Any]],
        format_enum: ExportFormat,
        output_path: Optional[str],
        options: ExportOptions
    ) -> Union[str, bytes]:
        """
        导出为 Excel 格式
        
        Args:
            assets: 资产数据列表
            format_enum: 格式枚举
            output_path: 输出路径
            options: 导出选项
            
        Returns:
            文件路径或字节数据
        """
        if self._engine == "openpyxl":
            return self._export_with_openpyxl(assets, format_enum, output_path, options)
        elif self._engine == "xlsxwriter":
            return self._export_with_xlsxwriter(assets, format_enum, output_path, options)
        else:
            raise RuntimeError("Excel engine not available")

    def _export_with_openpyxl(
        self,
        assets: List[Dict[str, Any]],
        format_enum: ExportFormat,
        output_path: Optional[str],
        options: ExportOptions
    ) -> Union[str, bytes]:
        """
        使用 openpyxl 导出数据
        
        Args:
            assets: 资产数据列表
            format_enum: 格式枚举
            output_path: 输出路径
            options: 导出选项
            
        Returns:
            文件路径或字节数据
        """
        wb = Workbook()
        ws = wb.active
        ws.title = options.sheet_name
        
        # 写入表头
        headers = [field.display_name for field in self.ASSET_FIELDS]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # 写入数据行
        for row_idx, asset in enumerate(assets, start=2):
            for col_idx, field in enumerate(self.ASSET_FIELDS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = self._format_field_value(asset.get(field.name), field, options)
                cell.value = value
                cell.font = self.CELL_FONT
                cell.alignment = self.CELL_ALIGNMENT
                cell.border = self.THIN_BORDER
        
        # 自适应列宽
        for col_idx, field in enumerate(self.ASSET_FIELDS, start=1):
            max_length = len(field.display_name)
            for row_idx in range(2, len(assets) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # 保存文件
        if output_path:
            wb.save(output_path)
            logger.info(f"Export saved to: {output_path}")
            return output_path
        else:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

    def _export_with_xlsxwriter(
        self,
        assets: List[Dict[str, Any]],
        format_enum: ExportFormat,
        output_path: Optional[str],
        options: ExportOptions
    ) -> Union[str, bytes]:
        """
        使用 xlsxwriter 导出数据
        
        Args:
            assets: 资产数据列表
            format_enum: 格式枚举
            output_path: 输出路径
            options: 导出选项
            
        Returns:
            文件路径或字节数据
        """
        if not output_path:
            raise ValueError("xlsxwriter requires output_path")
        
        wb = xlsxwriter.Workbook(output_path)
        ws = wb.add_worksheet(options.sheet_name)
        
        # 定义格式
        header_format = wb.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '4472C4',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        cell_format = wb.add_format({
            'align': 'left',
            'valign': 'vcenter',
            'border': 1
        })
        
        # 写入表头
        headers = [field.display_name for field in self.ASSET_FIELDS]
        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header, header_format)
        
        # 写入数据
        for row_idx, asset in enumerate(assets, start=1):
            for col_idx, field in enumerate(self.ASSET_FIELDS):
                value = self._format_field_value(asset.get(field.name), field, options)
                ws.write(row_idx, col_idx, value, cell_format)
        
        wb.close()
        logger.info(f"Export saved to: {output_path}")
        return output_path

    def _export_csv(
        self,
        assets: List[Dict[str, Any]],
        output_path: Optional[str],
        options: ExportOptions
    ) -> Union[str, bytes]:
        """
        导出为 CSV 格式
        
        Args:
            assets: 资产数据列表
            output_path: 输出路径
            options: 导出选项
            
        Returns:
            文件路径或 CSV 字符串
        """
        if output_path:
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                self._write_csv_export(writer, assets, options)
            return output_path
        else:
            output = io.StringIO()
            output.write('\ufeff')  # BOM for Excel
            writer = csv.writer(output)
            self._write_csv_export(writer, assets, options)
            return output.getvalue()

    def _write_csv_export(self, writer, assets: List[Dict[str, Any]], options: ExportOptions) -> None:
        """
        写入 CSV 导出内容
        
        Args:
            writer: CSV writer 对象
            assets: 资产数据列表
            options: 导出选项
        """
        # 写入表头
        if options.include_header:
            headers = [field.display_name for field in self.ASSET_FIELDS]
            writer.writerow(headers)
        
        # 写入数据行
        for asset in assets:
            row = []
            for field in self.ASSET_FIELDS:
                value = self._format_field_value(asset.get(field.name), field, options)
                row.append(value)
            writer.writerow(row)

    def _format_field_value(
        self,
        value: Any,
        field: AssetField,
        options: ExportOptions
    ) -> Any:
        """
        格式化字段值
        
        根据字段类型和导出选项格式化值：
        - 日期类型：格式化为指定格式
        - 金额类型：保留指定小数位
        - 枚举类型：保持原值
        
        Args:
            value: 原始值
            field: 字段定义
            options: 导出选项
            
        Returns:
            格式化后的值
        """
        if value is None:
            return ""
        
        if field.validation_type == "date":
            if isinstance(value, (datetime, date)):
                return value.strftime(options.date_format)
            elif isinstance(value, str):
                # 尝试解析日期字符串
                try:
                    dt = datetime.fromisoformat(value.replace('/', '-'))
                    return dt.strftime(options.date_format)
                except ValueError:
                    return value
        
        elif field.validation_type == "number":
            if isinstance(value, (int, float, Decimal)):
                return round(float(value), options.decimal_places)
        
        return value

    def generate_error_report(
        self,
        errors: List[ImportError],
        format_type: Union[str, ExportFormat] = ExportFormat.XLSX,
        output_path: Optional[str] = None,
        sheet_name: str = "导入错误报告"
    ) -> Union[str, bytes]:
        """
        生成导入错误报告
        
        将导入过程中的错误信息整理为可读的报告文件
        
        Args:
            errors: 错误信息列表
            format_type: 导出格式，支持 'xlsx', 'xls', 'csv'
            output_path: 文件保存路径，为 None 时返回字节数据
            sheet_name: 工作表名称（仅 Excel 格式有效）
            
        Returns:
            str: 当 output_path 指定时返回文件路径
            bytes: 当 output_path 为 None 时返回文件内容
            
        典型用法:
            >>> generator = ExcelGenerator()
            >>> errors = [
            ...     ImportError(row=3, field="asset_id", message="资产编号重复"),
            ...     ImportError(row=7, field="purchase_date", message="日期格式错误"),
            ... ]
            >>> report = generator.generate_error_report(errors, "xlsx", "/tmp/error_report.xlsx")
        """
        if not errors:
            logger.warning("No errors to report")
            return output_path or ""
        
        format_enum = self._validate_format(format_type)
        
        logger.info(f"Generating error report: {len(errors)} errors, format={format_enum.value}")
        
        if format_enum == ExportFormat.CSV:
            return self._generate_error_report_csv(errors, output_path)
        else:
            return self._generate_error_report_excel(errors, format_enum, output_path, sheet_name)

    def _generate_error_report_excel(
        self,
        errors: List[ImportError],
        format_enum: ExportFormat,
        output_path: Optional[str],
        sheet_name: str
    ) -> Union[str, bytes]:
        """
        生成 Excel 格式的错误报告
        
        Args:
            errors: 错误信息列表
            format_enum: 格式枚举
            output_path: 输出路径
            sheet_name: 工作表名称
            
        Returns:
            文件路径或字节数据
        """
        if self._engine == "openpyxl":
            return self._generate_error_excel_openpyxl(errors, output_path, sheet_name)
        elif self._engine == "xlsxwriter":
            return self._generate_error_excel_xlsxwriter(errors, output_path, sheet_name)
        else:
            raise RuntimeError("Excel engine not available")

    def _generate_error_excel_openpyxl(
        self,
        errors: List[ImportError],
        output_path: Optional[str],
        sheet_name: str
    ) -> Union[str, bytes]:
        """
        使用 openpyxl 生成错误报告
        
        Args:
            errors: 错误信息列表
            output_path: 输出路径
            sheet_name: 工作表名称
            
        Returns:
            文件路径或字节数据
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 表头样式
        error_header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        # 写入表头
        headers = ["行号", "字段", "错误信息", "原始值"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
            cell.fill = error_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER
        
        # 设置列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30
        
        # 写入错误数据
        error_fill = PatternFill(start_color="FFE4E4", end_color="FFE4E4", fill_type="solid")
        for row_idx, error in enumerate(errors, start=2):
            ws.cell(row=row_idx, column=1).value = error.row
            ws.cell(row=row_idx, column=2).value = error.field
            ws.cell(row=row_idx, column=3).value = error.message
            ws.cell(row=row_idx, column=4).value = error.original_value or ""
            
            # 应用样式
            for col_idx in range(1, 5):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = error_fill
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 保存
        if output_path:
            wb.save(output_path)
            logger.info(f"Error report saved to: {output_path}")
            return output_path
        else:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

    def _generate_error_excel_xlsxwriter(
        self,
        errors: List[ImportError],
        output_path: str,
        sheet_name: str
    ) -> str:
        """
        使用 xlsxwriter 生成错误报告
        
        Args:
            errors: 错误信息列表
            output_path: 输出路径
            sheet_name: 工作表名称
            
        Returns:
            文件路径
        """
        wb = xlsxwriter.Workbook(output_path)
        ws = wb.add_worksheet(sheet_name)
        
        # 格式定义
        header_format = wb.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': 'C00000',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        error_format = wb.add_format({
            'bg_color': 'FFE4E4',
            'align': 'left',
            'valign': 'vcenter',
            'border': 1
        })
        
        # 写入表头
        headers = ["行号", "字段", "错误信息", "原始值"]
        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header, header_format)
        
        # 设置列宽
        ws.set_column(0, 0, 10)
        ws.set_column(1, 1, 20)
        ws.set_column(2, 2, 40)
        ws.set_column(3, 3, 30)
        
        # 写入错误数据
        for row_idx, error in enumerate(errors, start=1):
            ws.write(row_idx, 0, error.row, error_format)
            ws.write(row_idx, 1, error.field, error_format)
            ws.write(row_idx, 2, error.message, error_format)
            ws.write(row_idx, 3, error.original_value or "", error_format)
        
        wb.close()
        logger.info(f"Error report saved to: {output_path}")
        return output_path

    def _generate_error_report_csv(
        self,
        errors: List[ImportError],
        output_path: Optional[str]
    ) -> Union[str, bytes]:
        """
        生成 CSV 格式的错误报告
        
        Args:
            errors: 错误信息列表
            output_path: 输出路径
            
        Returns:
            文件路径或 CSV 字符串
        """
        if output_path:
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                # 写入表头
                writer.writerow(["行号", "字段", "错误信息", "原始值"])
                # 写入错误数据
                for error in errors:
                    writer.writerow([
                        error.row,
                        error.field,
                        error.message,
                        error.original_value or ""
                    ])
            return output_path
        else:
            output = io.StringIO()
            output.write('\ufeff')
            writer = csv.writer(output)
            writer.writerow(["行号", "字段", "错误信息", "原始值"])
            for error in errors:
                writer.writerow([
                    error.row,
                    error.field,
                    error.message,
                    error.original_value or ""
                ])
            return output.getvalue()

    def get_template_info(self) -> Dict[str, Any]:
        """
        获取模板信息摘要
        
        返回当前模板的字段定义和约束信息
        
        Returns:
            Dict[str, Any]: 模板信息字典
        """
        return {
            "format_support": [f.value for f in ExportFormat],
            "field_count": len(self.ASSET_FIELDS),
            "required_fields": [f.name for f in self.ASSET_FIELDS if f.required],
            "optional_fields": [f.name for f in self.ASSET_FIELDS if not f.required],
            "fields": [
                {
                    "name": f.name,
                    "display_name": f.display_name,
                    "required": f.required,
                    "validation_type": f.validation_type,
                    "allowed_values": f.allowed_values,
                    "max_length": f.max_length,
                    "description": f.description
                }
                for f in self.ASSET_FIELDS
            ]
        }


# 便捷函数接口
def generate_import_template(
    format_type: str = "xlsx",
    output_path: Optional[str] = None
) -> Union[str, bytes]:
    """
    便捷函数：生成资产导入模板
    
    Args:
        format_type: 导出格式
        output_path: 输出路径
        
    Returns:
        文件路径或字节数据
    """
    generator = ExcelGenerator()
    return generator.generate_import_template(format_type, output_path)


def export_assets(
    assets: List[Dict[str, Any]],
    format_type: str = "xlsx",
    output_path: Optional[str] = None
) -> Union[str, bytes]:
    """
    便捷函数：批量导出资产数据
    
    Args:
        assets: 资产数据列表
        format_type: 导出格式
        output_path: 输出路径
        
    Returns:
        文件路径或字节数据
    """
    generator = ExcelGenerator()
    return generator.export_assets(assets, format_type, output_path)


def generate_error_report(
    errors: List[Dict[str, Any]],
    format_type: str = "xlsx",
    output_path: Optional[str] = None
) -> Union[str, bytes]:
    """
    便捷函数：生成导入错误报告
    
    Args:
        errors: 错误信息列表
        format_type: 导出格式
        output_path: 输出路径
        
    Returns:
        文件路径或字节数据
    """
    generator = ExcelGenerator()
    import_errors = [
        ImportError(
            row=e.get("row", 0),
            field=e.get("field", ""),
            message=e.get("message", ""),
            original_value=e.get("original_value")
        )
        for e in errors
    ]
    return generator.generate_error_report(import_errors, format_type, output_path)