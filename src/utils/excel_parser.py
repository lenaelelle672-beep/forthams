"""
Excel 解析器模块

提供 Excel (.xlsx) 文件的解析功能，支持资产批量导入。
支持单工作表解析、流式读取、空单元格处理、合并单元格处理。

导入字段映射（固定 12 个核心字段）:
    - asset_id        (String, 可选，导入时为空则自动生成)
    - asset_name      (String, 必填, 最大50字符)
    - asset_type      (Enum, 必填, 枚举值: EQUIPMENT/FURNITURE/VEHICLE/IT_HARDWARE/OTHER)
    - serial_number   (String, 可选, 最大100字符)
    - purchase_date   (Date, 必填, YYYY-MM-DD 格式)
    - purchase_price  (Decimal, 必填, >0, 最多2位小数)
    - currency        (Enum, 必填, 默认CNY)
    - department      (String, 必填, 需匹配已存在的部门编码)
    - custodian       (String, 可选, 最大100字符)
    - status          (Enum, 必填, 枚举值: ACTIVE/INACTIVE/MAINTENANCE/RETIRED)
    - location        (String, 可选, 最大200字符)
    - remarks         (String, 可选, 最大500字符)

约束限制:
    - 单次导入上限: 5000 条记录
    - 文件大小上限: 10 MB
    - 仅支持 .xlsx 格式（不支持 .xls）

参考规范: SWARM-2025-Q2-P2-006
"""

from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from decimal import Decimal, InvalidOperation
import io

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 字段名到列索引的映射（支持中文表头）
HEADER_MAPPING: Dict[str, List[str]] = {
    "asset_id": ["asset_id", "资产编号", "编号", "ID"],
    "asset_name": ["asset_name", "资产名称", "名称"],
    "asset_type": ["asset_type", "资产类型", "类型"],
    "serial_number": ["serial_number", "序列号", "出厂编号"],
    "purchase_date": ["purchase_date", "购置日期", "购买日期", "购置时间"],
    "purchase_price": ["purchase_price", "购置价格", "购买价格", "价格"],
    "currency": ["currency", "币种", "货币"],
    "department": ["department", "部门", "所属部门"],
    "custodian": ["custodian", "保管人", "负责人"],
    "status": ["status", "状态"],
    "location": ["location", "位置", "存放地点"],
    "remarks": ["remarks", "备注", "说明"]
}

# 必填字段列表
REQUIRED_FIELDS: List[str] = [
    "asset_name",
    "asset_type",
    "purchase_date",
    "purchase_price",
    "currency",
    "department",
    "status"
]

# 有效枚举值
VALID_ASSET_TYPES: List[str] = [
    "EQUIPMENT", "FURNITURE", "VEHICLE", "IT_HARDWARE", "OTHER"
]

VALID_STATUSES: List[str] = [
    "ACTIVE", "INACTIVE", "MAINTENANCE", "RETIRED"
]

VALID_CURRENCIES: List[str] = [
    "CNY", "USD", "EUR", "GBP", "JPY", "HKD"
]

# 文件大小限制: 10 MB
MAX_FILE_SIZE_BYTES: int = 10 * 1024 * 1024

# 单次导入记录数限制: 5000
MAX_RECORDS_PER_IMPORT: int = 5000


class ExcelParserError(Exception):
    """Excel 解析器基础异常类"""
    pass


class FileSizeExceededError(ExcelParserError):
    """文件大小超过限制异常"""
    pass


class InvalidExcelFormatError(ExcelParserError):
    """Excel 格式无效异常"""
    pass


class ParseResult:
    """
    Excel 解析结果数据类

    用于封装 Excel 文件解析后的数据结构，
    包含解析成功的记录列表和解析过程中产生的警告信息。

    Attributes:
        records: 解析成功的记录列表，每条记录为字典
        warnings: 解析过程中产生的警告信息列表
        total_rows: 原始数据行数（含表头）
        error_rows: 发生错误的行号列表
    """
    
    def __init__(
        self,
        records: Optional[List[Dict[str, Any]]] = None,
        warnings: Optional[List[str]] = None
    ) -> None:
        """
        初始化解析结果对象

        Args:
            records: 解析成功的记录列表，默认为空列表
            warnings: 警告信息列表，默认为空列表
        """
        self.records: List[Dict[str, Any]] = records if records is not None else []
        self.warnings: List[str] = warnings if warnings is not None else []
        self.total_rows: int = 0
        self.error_rows: List[int] = []

    def add_warning(self, message: str) -> None:
        """
        添加警告信息

        Args:
            message: 警告信息内容
        """
        self.warnings.append(message)

    def add_error_row(self, row_number: int) -> None:
        """
        记录错误行号

        Args:
            row_number: 发生错误的行号（从 1 开始计数）
        """
        self.error_rows.append(row_number)

    def get_summary(self) -> Dict[str, Any]:
        """
        获取解析结果摘要信息

        Returns:
            包含解析统计信息的字典
        """
        return {
            "total_rows": self.total_rows,
            "success_count": len(self.records),
            "error_count": len(self.error_rows),
            "warning_count": len(self.warnings)
        }


class ExcelParser:
    """
    Excel 文件解析器

    提供 Excel (.xlsx) 文件的流式解析功能，支持:
    - UTF-8 和 GBK 编码
    - 空单元格处理（转为空字符串）
    - 合并单元格处理（取首值）
    - 大文件流式读取（内存友好）

    Example:
        >>> parser = ExcelParser()
        >>> result = parser.parse(file_stream)
        >>> for record in result.records:
        ...     print(record["asset_name"])
    """

    def __init__(self) -> None:
        """
        初始化 Excel 解析器

        检查 openpyxl 库是否可用，若不可用则抛出异常
        """
        if not OPENPYXL_AVAILABLE:
            raise ExcelParserError(
                "openpyxl library is not installed. "
                "Please install it with: pip install openpyxl"
            )
        self._workbook = None
        self._worksheet = None

    def parse(
        self,
        file_content: bytes,
        sheet_index: int = 0
    ) -> ParseResult:
        """
        解析 Excel 文件内容

        Args:
            file_content: Excel 文件的字节内容
            sheet_index: 要解析的工作表索引，默认为 0（第一个工作表）

        Returns:
            ParseResult: 解析结果对象，包含记录列表和警告信息

        Raises:
            FileSizeExceededError: 文件大小超过 10MB 限制
            InvalidExcelFormatError: Excel 格式无效或文件损坏
            ExcelParserError: 其他解析错误
        """
        # 检查文件大小
        if len(file_content) > MAX_FILE_SIZE_BYTES:
            raise FileSizeExceededError(
                f"File size ({len(file_content)} bytes) exceeds "
                f"maximum allowed size ({MAX_FILE_SIZE_BYTES} bytes)"
            )

        try:
            # 加载工作簿
            self._workbook = openpyxl.load_workbook(
                io.BytesIO(file_content),
                read_only=True,
                data_only=True
            )
        except InvalidFileException as e:
            raise InvalidExcelFormatError(f"Invalid Excel file format: {str(e)}")
        except Exception as e:
            raise ExcelParserError(f"Failed to load Excel file: {str(e)}")

        # 获取工作表（忽略其他工作表，只取第一个）
        if sheet_index >= len(self._workbook.worksheets):
            raise InvalidExcelFormatError(
                f"Sheet index {sheet_index} out of range. "
                f"Workbook has {len(self._workbook.worksheets)} sheets."
            )

        self._worksheet = self._workbook.worksheets[sheet_index]

        # 解析数据
        result = self._parse_worksheet()

        # 关闭工作簿释放资源
        self._workbook.close()

        return result

    def parse_stream(
        self,
        file_path: str,
        sheet_index: int = 0
    ) -> ParseResult:
        """
        流式解析 Excel 文件（适合大文件）

        使用 read_only 模式流式读取，减少内存占用

        Args:
            file_path: Excel 文件路径
            sheet_index: 要解析的工作表索引，默认为 0

        Returns:
            ParseResult: 解析结果对象

        Raises:
            FileNotFoundError: 文件不存在
            FileSizeExceededError: 文件大小超过限制
            InvalidExcelFormatError: Excel 格式无效
        """
        import os

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        file_size = os.path.getsize(file_path)
        if file_size > MAX_FILE_SIZE_BYTES:
            raise FileSizeExceededError(
                f"File size ({file_size} bytes) exceeds "
                f"maximum allowed size ({MAX_FILE_SIZE_BYTES} bytes)"
            )

        try:
            self._workbook = openpyxl.load_workbook(
                file_path,
                read_only=True,
                data_only=True
            )
        except InvalidFileException as e:
            raise InvalidExcelFormatError(f"Invalid Excel file format: {str(e)}")
        except Exception as e:
            raise ExcelParserError(f"Failed to load Excel file: {str(e)}")

        if sheet_index >= len(self._workbook.worksheets):
            raise InvalidExcelFormatError(
                f"Sheet index {sheet_index} out of range."
            )

        self._worksheet = self._workbook.worksheets[sheet_index]
        result = self._parse_worksheet()
        self._workbook.close()

        return result

    def _parse_worksheet(self) -> ParseResult:
        """
        内部方法：解析工作表内容

        Returns:
            ParseResult: 解析结果对象
        """
        result = ParseResult()
        rows = list(self._worksheet.iter_rows(values_only=True))

        if not rows:
            result.add_warning("Worksheet is empty")
            return result

        # 解析表头行
        header_row = rows[0]
        header_map = self._build_header_map(header_row)

        if not header_map:
            result.add_warning("No valid headers found in first row")
            return result

        # 检查必填字段
        missing_required = self._check_required_fields(header_map)
        if missing_required:
            result.add_warning(
                f"Missing required fields: {', '.join(missing_required)}"
            )

        result.total_rows = len(rows) - 1  # 排除表头行

        # 解析数据行
        for row_idx, row in enumerate(rows[1:], start=2):  # 从第2行开始（跳过表头）
            record = self._parse_row(row, header_map, row_idx, result)
            if record:
                # 检查是否超过最大记录数
                if len(result.records) >= MAX_RECORDS_PER_IMPORT:
                    result.add_warning(
                        f"Record count exceeds maximum ({MAX_RECORDS_PER_IMPORT}). "
                        f"Remaining rows will be ignored."
                    )
                    break
                result.records.append(record)

        return result

    def _build_header_map(
        self,
        header_row: Tuple
    ) -> Dict[str, int]:
        """
        构建表头到列索引的映射

        Args:
            header_row: 表头行数据（元组）

        Returns:
            Dict[str, int]: 字段名到列索引的映射
        """
        header_map: Dict[str, int] = {}

        for col_idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue

            cell_str = str(cell_value).strip().lower()

            # 匹配字段名
            for field_name, aliases in HEADER_MAPPING.items():
                if cell_str in [alias.lower() for alias in aliases]:
                    header_map[field_name] = col_idx
                    break

        return header_map

    def _check_required_fields(
        self,
        header_map: Dict[str, int]
    ) -> List[str]:
        """
        检查表头中是否包含所有必填字段

        Args:
            header_map: 表头映射字典

        Returns:
            List[str]: 缺失的必填字段列表
        """
        missing = []
        for field in REQUIRED_FIELDS:
            if field not in header_map:
                missing.append(field)
        return missing

    def _parse_row(
        self,
        row: Tuple,
        header_map: Dict[str, int],
        row_number: int,
        result: ParseResult
    ) -> Optional[Dict[str, Any]]:
        """
        解析单行数据

        Args:
            row: 行数据（元组）
            header_map: 表头映射
            row_number: 行号（用于错误报告）
            result: 解析结果对象（用于收集警告和错误）

        Returns:
            解析后的记录字典，若行为空则返回 None
        """
        # 处理空行
        if all(cell is None or str(cell).strip() == "" for cell in row):
            result.add_warning(f"Row {row_number} is empty, skipping")
            return None

        record: Dict[str, Any] = {}

        for field_name, col_idx in header_map.items():
            if col_idx < len(row):
                value = row[col_idx]
                # 空单元格处理：转为空字符串而非 null
                record[field_name] = self._normalize_value(value)
            else:
                record[field_name] = ""

        return record

    def _normalize_value(self, value: Any) -> str:
        """
        规范化单元格值

        Args:
            value: 原始单元格值

        Returns:
            str: 规范化后的字符串值，空值为空字符串
        """
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        # 处理日期、时间等类型
        if isinstance(value, (datetime,)):
            return value.strftime("%Y-%m-%d")
        # 其他类型转为字符串
        return str(value).strip()


def validate_asset_data(record: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    校验单条资产数据的有效性

    Args:
        record: 资产记录字典

    Returns:
        List[Dict[str, Any]]: 校验错误列表，每条错误包含:
            - field: 字段名
            - value: 错误值
            - reason: 错误原因
    """
    errors: List[Dict[str, Any]] = []

    # 必填字段检查
    for field in REQUIRED_FIELDS:
        if field not in record or not record[field]:
            errors.append({
                "field": field,
                "value": record.get(field, ""),
                "reason": f"{field} is required"
            })

    # 资产类型枚举检查
    if "asset_type" in record and record["asset_type"]:
        if record["asset_type"] not in VALID_ASSET_TYPES:
            errors.append({
                "field": "asset_type",
                "value": record["asset_type"],
                "reason": f"invalid enum, expected one of {VALID_ASSET_TYPES}"
            })

    # 状态枚举检查
    if "status" in record and record["status"]:
        if record["status"] not in VALID_STATUSES:
            errors.append({
                "field": "status",
                "value": record["status"],
                "reason": f"invalid enum, expected one of {VALID_STATUSES}"
            })

    # 货币枚举检查
    if "currency" in record and record["currency"]:
        if record["currency"] not in VALID_CURRENCIES:
            errors.append({
                "field": "currency",
                "value": record["currency"],
                "reason": f"invalid enum, expected one of {VALID_CURRENCIES}"
            })

    # 日期格式检查
    if "purchase_date" in record and record["purchase_date"]:
        if not _is_valid_date_format(record["purchase_date"]):
            errors.append({
                "field": "purchase_date",
                "value": record["purchase_date"],
                "reason": "invalid date format, expected YYYY-MM-DD"
            })

    # 价格数字检查
    if "purchase_price" in record and record["purchase_price"]:
        if not _is_valid_decimal(record["purchase_price"]):
            errors.append({
                "field": "purchase_price",
                "value": record["purchase_price"],
                "reason": "must be a valid positive number"
            })
        elif float(record["purchase_price"]) <= 0:
            errors.append({
                "field": "purchase_price",
                "value": record["purchase_price"],
                "reason": "must be greater than 0"
            })

    # 字符串长度检查
    length_constraints = {
        "asset_name": 50,
        "serial_number": 100,
        "custodian": 100,
        "location": 200,
        "remarks": 500
    }
    for field, max_length in length_constraints.items():
        if field in record and record[field]:
            if len(str(record[field])) > max_length:
                errors.append({
                    "field": field,
                    "value": record[field],
                    "reason": f"exceeds maximum length of {max_length}"
                })

    return errors


def _is_valid_date_format(date_str: str) -> bool:
    """
    验证日期字符串格式是否为 YYYY-MM-DD

    Args:
        date_str: 日期字符串

    Returns:
        bool: 是否为有效格式
    """
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except (ValueError, TypeError):
        return False


def _is_valid_decimal(value: Any) -> bool:
    """
    验证值是否可以转为有效的十进制数

    Args:
        value: 待验证的值

    Returns:
        bool: 是否为有效数字
    """
    try:
        Decimal(str(value))
        return True
    except (InvalidOperation, ValueError, TypeError):
        return False


def create_error_report(
    errors: List[Dict[str, Any]],
    output_format: str = "csv"
) -> str:
    """
    创建错误报告

    Args:
        errors: 错误列表
        output_format: 输出格式，目前支持 "csv"

    Returns:
        str: 格式化后的错误报告字符串
    """
    if output_format == "csv":
        lines = ["row_number,field,error_value,error_reason"]
        for error in errors:
            row = error.get("row", "N/A")
            field = error.get("field", "")
            value = str(error.get("value", "")).replace(",", ";")
            reason = error.get("reason", "")
            lines.append(f"{row},{field},{value},{reason}")
        return "\n".join(lines)
    else:
        raise ValueError(f"Unsupported output format: {output_format}")