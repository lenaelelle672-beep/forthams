"""
资产批量导入导出 - 文件解析模块

提供 Excel/CSV 文件的统一解析接口，支持字段映射和基础校验。

支持的格式：
    - Excel: .xlsx, .xls
    - CSV: .csv (UTF-8 编码)

功能特性：
    - 字段级校验规则（数据类型、长度、枚举值）
    - 部分导入模式（跳过/标记错误行继续处理）
    - 大文件流式解析（异步处理机制）

使用示例：
    >>> from src.utils.file_parser import FileParser, FileParseError
    >>> parser = FileParser()
    >>> result = parser.parse("assets.xlsx")
    >>> print(f"成功解析 {result['success_count']} 行")
"""

import logging
import csv
import io
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from enum import Enum
from typing import Any, BinaryIO, Optional

logger = logging.getLogger(__name__)


class AssetTypeEnum(str, Enum):
    """资产类型枚举"""
    EQUIPMENT = "EQUIPMENT"
    INSTRUMENT = "INSTRUMENT"
    VEHICLE = "VEHICLE"
    OTHER = "OTHER"


class AssetStatusEnum(str, Enum):
    """资产状态枚举"""
    ACTIVE = "ACTIVE"
    MAINTENANCE = "MAINTENANCE"
    SCRAPPED = "SCRAPPED"


class FileParseError(Exception):
    """文件解析错误异常"""
    
    def __init__(self, message: str, row: Optional[int] = None, field: Optional[str] = None):
        """
        初始化解析错误
        
        Args:
            message: 错误描述信息
            row: 错误发生的行号（从1开始）
            field: 错误发生的字段名
        """
        self.message = message
        self.row = row
        self.field = field
        super().__init__(self._format_message())
    
    def _format_message(self) -> str:
        """格式化错误消息"""
        parts = [self.message]
        if self.row is not None:
            parts.append(f"行号: {self.row}")
        if self.field is not None:
            parts.append(f"字段: {self.field}")
        return ", ".join(parts)


class ValidationError(Exception):
    """数据校验错误异常"""
    
    def __init__(self, field: str, message: str, row: int):
        """
        初始化校验错误
        
        Args:
            field: 字段名
            message: 校验失败原因
            row: 行号
        """
        self.field = field
        self.message = message
        self.row = row
        super().__init__(f"[行{row}] {field}: {message}")


@dataclass
class ParseResult:
    """
    文件解析结果数据类
    
    Attributes:
        success_count: 成功解析的行数
        failed_count: 解析失败的行数
        errors: 错误列表
        data: 解析后的数据列表
        total_rows: 总行数
    """
    success_count: int = 0
    failed_count: int = 0
    errors: list = field(default_factory=list)
    data: list = field(default_factory=list)
    total_rows: int = 0


@dataclass
class AssetRowData:
    """
    资产行数据类
    
    表示一条解析后的资产记录，包含所有字段的校验后值。
    """
    asset_id: str
    asset_name: str
    asset_type: AssetTypeEnum
    purchase_date: datetime
    purchase_amount: Decimal
    department: str
    status: AssetStatusEnum
    location: Optional[str] = None
    description: Optional[str] = None
    row_number: int = 0
    is_valid: bool = True
    validation_errors: list = field(default_factory=list)


class BaseParser(ABC):
    """文件解析器抽象基类"""
    
    REQUIRED_FIELDS = [
        "asset_id", "asset_name", "asset_type", "purchase_date",
        "purchase_amount", "department", "status"
    ]
    
    OPTIONAL_FIELDS = ["location", "description"]
    
    FIELD_MAPPING = {
        "资产编号": "asset_id",
        "资产名称": "asset_name",
        "资产类型": "asset_type",
        "购置日期": "purchase_date",
        "购置金额": "purchase_amount",
        "使用部门": "department",
        "资产状态": "status",
        "存放地点": "location",
        "备注": "description"
    }
    
    def __init__(self, mode: str = "partial"):
        """
        初始化解析器
        
        Args:
            mode: 解析模式
                - "strict": 全量失败模式，遇到错误立即停止
                - "partial": 部分导入模式，跳过错误行继续处理
        """
        self.mode = mode
    
    @abstractmethod
    def parse(self, file_content: BinaryIO) -> ParseResult:
        """
        解析文件内容
        
        Args:
            file_content: 文件二进制内容
            
        Returns:
            ParseResult: 解析结果对象
        """
        pass
    
    @abstractmethod
    def _read_rows(self, file_content: BinaryIO) -> list:
        """
        读取文件行数据
        
        Args:
            file_content: 文件二进制内容
            
        Returns:
            list: 行数据列表
        """
        pass
    
    def _map_headers(self, headers: list) -> dict:
        """
        映射表头字段名
        
        Args:
            headers: 原始表头列表
            
        Returns:
            dict: 字段名映射后的表头字典
        """
        mapped = {}
        for idx, header in enumerate(headers):
            normalized = header.strip()
            mapped[idx] = self.FIELD_MAPPING.get(normalized, normalized)
        return mapped
    
    def _validate_asset_id(self, value: str, row: int) -> tuple:
        """
        校验资产编号
        
        Args:
            value: 资产编号值
            row: 行号
            
        Returns:
            tuple: (校验结果, 错误信息)
        """
        if not value or not value.strip():
            return False, "资产编号不能为空"
        
        cleaned = value.strip()
        if len(cleaned) > 64:
            return False, f"资产编号长度不能超过64字符，当前: {len(cleaned)}"
        
        if not all(c.isalnum() or c == '_' for c in cleaned):
            return False, "资产编号只能包含字母、数字和下划线"
        
        return True, None
    
    def _validate_asset_name(self, value: str, row: int) -> tuple:
        """
        校验资产名称
        
        Args:
            value: 资产名称值
            row: 行号
            
        Returns:
            tuple: (校验结果, 错误信息)
        """
        if not value or not value.strip():
            return False, "资产名称不能为空"
        
        cleaned = value.strip()
        if len(cleaned) > 128:
            return False, f"资产名称长度不能超过128字符，当前: {len(cleaned)}"
        
        return True, None
    
    def _validate_asset_type(self, value: str, row: int) -> tuple:
        """
        校验资产类型枚举值
        
        Args:
            value: 资产类型值
            row: 行号
            
        Returns:
            tuple: (校验结果, 错误信息)
        """
        if not value or not value.strip():
            return False, "资产类型不能为空"
        
        try:
            AssetTypeEnum(value.strip().upper())
            return True, None
        except ValueError:
            valid_types = [e.value for e in AssetTypeEnum]
            return False, f"资产类型枚举值不匹配，可选值: {valid_types}"
    
    def _validate_purchase_date(self, value: str, row: int) -> tuple:
        """
        校验购置日期格式
        
        Args:
            value: 购置日期值
            row: 行号
            
        Returns:
            tuple: (校验结果, 解析后的日期对象或错误信息)
        """
        if not value or not value.strip():
            return False, "购置日期不能为空"
        
        date_formats = ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"]
        cleaned = value.strip()
        
        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(cleaned, fmt)
                return True, parsed_date
            except ValueError:
                continue
        
        return False, f"日期格式错误，请使用 ISO 8601 格式 (YYYY-MM-DD)，当前值: {cleaned}"
    
    def _validate_purchase_amount(self, value: Any, row: int) -> tuple:
        """
        校验购置金额
        
        Args:
            value: 购置金额值
            row: 行号
            
        Returns:
            tuple: (校验结果, 解析后的金额对象或错误信息)
        """
        if value is None or (isinstance(value, str) and not value.strip()):
            return False, "购置金额不能为空"
        
        try:
            # 处理字符串类型的金额
            if isinstance(value, str):
                # 移除可能的货币符号和千分位逗号
                cleaned = value.strip().replace(",", "").replace("¥", "").replace("$", "")
                amount = Decimal(cleaned)
            else:
                amount = Decimal(str(value))
            
            if amount < 0:
                return False, "购置金额必须大于等于0"
            
            # 精度限制为2位小数
            amount = amount.quantize(Decimal("0.01"))
            return True, amount
            
        except InvalidOperation:
            return False, f"购置金额格式错误: {value}"
    
    def _validate_department(self, value: str, row: int, valid_departments: Optional[set] = None) -> tuple:
        """
        校验使用部门
        
        Args:
            value: 部门值
            row: 行号
            valid_departments: 有效部门编码集合，用于校验部门是否存在
            
        Returns:
            tuple: (校验结果, 错误信息)
        """
        if not value or not value.strip():
            return False, "使用部门不能为空"
        
        cleaned = value.strip()
        if len(cleaned) > 64:
            return False, f"部门编码长度不能超过64字符，当前: {len(cleaned)}"
        
        if valid_departments and cleaned not in valid_departments:
            return False, f"部门不存在: {cleaned}"
        
        return True, None
    
    def _validate_status(self, value: str, row: int) -> tuple:
        """
        校验资产状态枚举值
        
        Args:
            value: 状态值
            row: 行号
            
        Returns:
            tuple: (校验结果, 错误信息)
        """
        if not value or not value.strip():
            return False, "资产状态不能为空"
        
        try:
            AssetStatusEnum(value.strip().upper())
            return True, None
        except ValueError:
            valid_statuses = [e.value for e in AssetStatusEnum]
            return False, f"资产状态枚举值不匹配，可选值: {valid_statuses}"
    
    def _validate_field(self, field_name: str, value: Any, row: int, 
                       valid_departments: Optional[set] = None) -> tuple:
        """
        统一字段校验入口
        
        Args:
            field_name: 字段名
            value: 字段值
            row: 行号
            valid_departments: 有效部门集合
            
        Returns:
            tuple: (校验结果, 校验后的值或错误信息)
        """
        validators = {
            "asset_id": self._validate_asset_id,
            "asset_name": self._validate_asset_name,
            "asset_type": self._validate_asset_type,
            "purchase_date": self._validate_purchase_date,
            "purchase_amount": self._validate_purchase_amount,
            "department": lambda v, r: self._validate_department(v, r, valid_departments),
            "status": self._validate_status,
        }
        
        validator = validators.get(field_name)
        if validator:
            return validator(value, row)
        
        return True, value
    
    def validate_row(self, row_data: dict, row_number: int,
                   valid_departments: Optional[set] = None) -> AssetRowData:
        """
        校验单行数据
        
        Args:
            row_data: 行数据字典
            row_number: 行号
            valid_departments: 有效部门集合
            
        Returns:
            AssetRowData: 校验后的资产行数据对象
        """
        errors = []
        validated_data = {}
        
        for field_name in self.REQUIRED_FIELDS:
            value = row_data.get(field_name)
            is_valid, result = self._validate_field(field_name, value, row_number, valid_departments)
            
            if not is_valid:
                errors.append(ValidationError(field_name, result, row_number))
            else:
                validated_data[field_name] = result
        
        # 处理可选字段
        for field_name in self.OPTIONAL_FIELDS:
            value = row_data.get(field_name)
            if value and value.strip():
                validated_data[field_name] = value.strip()[:128 if field_name == "location" else 512]
        
        asset_data = AssetRowData(
            asset_id=validated_data.get("asset_id", ""),
            asset_name=validated_data.get("asset_name", ""),
            asset_type=validated_data.get("asset_type", AssetTypeEnum.OTHER),
            purchase_date=validated_data.get("purchase_date", datetime.now()),
            purchase_amount=validated_data.get("purchase_amount", Decimal("0")),
            department=validated_data.get("department", ""),
            status=validated_data.get("status", AssetStatusEnum.ACTIVE),
            location=validated_data.get("location"),
            description=validated_data.get("description"),
            row_number=row_number,
            is_valid=len(errors) == 0,
            validation_errors=errors
        )
        
        return asset_data


class CsvParser(BaseParser):
    """CSV 文件解析器"""
    
    def __init__(self, mode: str = "partial", encoding: str = "utf-8"):
        """
        初始化 CSV 解析器
        
        Args:
            mode: 解析模式
            encoding: 文件编码，默认 UTF-8
        """
        super().__init__(mode)
        self.encoding = encoding
    
    def _read_rows(self, file_content: BinaryIO) -> list:
        """
        读取 CSV 文件行数据
        
        Args:
            file_content: 文件二进制内容
            
        Returns:
            list: 行数据列表
        """
        content = file_content.read()
        
        # 检测并处理 BOM
        if content.startswith(b'\xef\xbb\xbf'):
            content = content[3:]
            self.encoding = "utf-8-sig"
        
        text = content.decode(self.encoding, errors="replace")
        reader = csv.reader(io.StringIO(text))
        
        rows = list(reader)
        if not rows:
            return []
        
        return rows
    
    def parse(self, file_content: BinaryIO, valid_departments: Optional[set] = None) -> ParseResult:
        """
        解析 CSV 文件内容
        
        Args:
            file_content: CSV 文件二进制内容
            valid_departments: 有效部门集合
            
        Returns:
            ParseResult: 解析结果对象
        """
        result = ParseResult()
        
        try:
            rows = self._read_rows(file_content)
            
            if not rows:
                raise FileParseError("CSV 文件为空或格式错误")
            
            # 第一行为表头
            header_row = rows[0]
            header_mapping = self._map_headers(header_row)
            data_rows = rows[1:]
            
            result.total_rows = len(data_rows)
            seen_ids = set()
            
            for idx, row in enumerate(data_rows):
                row_number = idx + 2  # 行号从2开始（1是表头）
                
                if self.mode == "strict" and result.failed_count > 0:
                    break
                
                # 构建行数据字典
                row_data = {}
                for col_idx, value in enumerate(row):
                    field_name = header_mapping.get(col_idx)
                    if field_name:
                        row_data[field_name] = value
                
                # 校验唯一性
                asset_id = row_data.get("asset_id", "").strip() if row_data.get("asset_id") else ""
                if asset_id:
                    if asset_id in seen_ids:
                        result.errors.append({
                            "row": row_number,
                            "field": "asset_id",
                            "message": f"重复的资产编号: {asset_id}"
                        })
                        result.failed_count += 1
                        continue
                    seen_ids.add(asset_id)
                
                # 校验行数据
                asset_data = self.validate_row(row_data, row_number, valid_departments)
                
                if asset_data.is_valid:
                    result.success_count += 1
                    result.data.append(asset_data)
                else:
                    result.failed_count += 1
                    for error in asset_data.validation_errors:
                        result.errors.append({
                            "row": error.row,
                            "field": error.field,
                            "message": error.message
                        })
            
            logger.info(
                f"CSV 解析完成: 总行数={result.total_rows}, "
                f"成功={result.success_count}, 失败={result.failed_count}"
            )
            
        except FileParseError:
            raise
        except Exception as e:
            logger.error(f"CSV 解析异常: {str(e)}")
            raise FileParseError(f"CSV 解析失败: {str(e)}")
        
        return result


class ExcelParser(BaseParser):
    """Excel 文件解析器"""
    
    def __init__(self, mode: str = "partial"):
        """
        初始化 Excel 解析器
        
        Args:
            mode: 解析模式
        """
        super().__init__(mode)
        self._openpyxl_available = self._check_openpyxl()
        self._xlrd_available = self._check_xlrd()
    
    def _check_openpyxl(self) -> bool:
        """检查 openpyxl 是否可用"""
        try:
            import openpyxl
            return True
        except ImportError:
            return False
    
    def _check_xlrd(self) -> bool:
        """检查 xlrd 是否可用"""
        try:
            import xlrd
            return True
        except ImportError:
            return False
    
    def _read_rows(self, file_content: BinaryIO) -> list:
        """
        读取 Excel 文件行数据
        
        Args:
            file_content: 文件二进制内容
            
        Returns:
            list: 行数据列表
        """
        if not self._openpyxl_available and not self._xlrd_available:
            raise FileParseError("未安装 Excel 解析库，请运行: pip install openpyxl xlrd")
        
        content = file_content.read()
        
        # 检测文件格式
        if content[:4] == b"PK\x03\x04":
            return self._read_xlsx(content)
        elif content[:8] == b"\xd0\xcf\x11\xe0":
            return self._read_xls(content)
        else:
            raise FileParseError("不支持的文件格式，请使用 .xlsx 或 .xls 文件")
    
    def _read_xlsx(self, content: bytes) -> list:
        """
        读取 xlsx 格式文件
        
        Args:
            content: 文件二进制内容
            
        Returns:
            list: 行数据列表
        """
        import openpyxl
        
        if not self._openpyxl_available:
            raise FileParseError("请安装 openpyxl: pip install openpyxl")
        
        from io import BytesIO
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active
        
        rows = []
        for row in sheet.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            rows.append(row_data)
        
        workbook.close()
        return rows
    
    def _read_xls(self, content: bytes) -> list:
        """
        读取 xls 格式文件
        
        Args:
            content: 文件二进制内容
            
        Returns:
            list: 行数据列表
        """
        if not self._xlrd_available:
            raise FileParseError("请安装 xlrd: pip install xlrd")
        
        import xlrd
        from io import BytesIO
        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        
        rows = []
        for row_idx in range(sheet.nrows):
            row_data = [str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
            rows.append(row_data)
        
        workbook.release_session()
        return rows
    
    def parse(self, file_content: BinaryIO, valid_departments: Optional[set] = None) -> ParseResult:
        """
        解析 Excel 文件内容
        
        Args:
            file_content: Excel 文件二进制内容
            valid_departments: 有效部门集合
            
        Returns:
            ParseResult: 解析结果对象
        """
        result = ParseResult()
        
        try:
            rows = self._read_rows(file_content)
            
            if not rows:
                raise FileParseError("Excel 文件为空或格式错误")
            
            # 第一行为表头
            header_row = rows[0]
            header_mapping = self._map_headers(header_row)
            data_rows = rows[1:]
            
            result.total_rows = len(data_rows)
            seen_ids = set()
            
            for idx, row in enumerate(data_rows):
                row_number = idx + 2  # 行号从2开始（1是表头）
                
                if self.mode == "strict" and result.failed_count > 0:
                    break
                
                # 构建行数据字典
                row_data = {}
                for col_idx, value in enumerate(row):
                    field_name = header_mapping.get(col_idx)
                    if field_name:
                        row_data[field_name] = value
                
                # 校验唯一性
                asset_id = row_data.get("asset_id", "").strip() if row_data.get("asset_id") else ""
                if asset_id:
                    if asset_id in seen_ids:
                        result.errors.append({
                            "row": row_number,
                            "field": "asset_id",
                            "message": f"重复的资产编号: {asset_id}"
                        })
                        result.failed_count += 1
                        continue
                    seen_ids.add(asset_id)
                
                # 校验行数据
                asset_data = self.validate_row(row_data, row_number, valid_departments)
                
                if asset_data.is_valid:
                    result.success_count += 1
                    result.data.append(asset_data)
                else:
                    result.failed_count += 1
                    for error in asset_data.validation_errors:
                        result.errors.append({
                            "row": error.row,
                            "field": error.field,
                            "message": error.message
                        })
            
            logger.info(
                f"Excel 解析完成: 总行数={result.total_rows}, "
                f"成功={result.success_count}, 失败={result.failed_count}"
            )
            
        except FileParseError:
            raise
        except Exception as e:
            logger.error(f"Excel 解析异常: {str(e)}")
            raise FileParseError(f"Excel 解析失败: {str(e)}")
        
        return result


class FileParser:
    """
    统一文件解析器
    
    根据文件扩展名自动选择合适的解析器，支持 Excel 和 CSV 格式。
    
    Attributes:
        mode: 解析模式 ("strict" 或 "partial")
        encoding: CSV 文件编码
    """
    
    SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
    
    def __init__(self, mode: str = "partial", encoding: str = "utf-8"):
        """
        初始化文件解析器
        
        Args:
            mode: 解析模式
                - "strict": 全量失败模式
                - "partial": 部分导入模式（默认）
            encoding: CSV 文件编码
        """
        self.mode = mode
        self.encoding = encoding
    
    def parse(self, file_content: BinaryIO, file_extension: str,
             valid_departments: Optional[set] = None) -> ParseResult:
        """
        解析文件内容
        
        Args:
            file_content: 文件二进制内容
            file_extension: 文件扩展名（包含点号，如 ".xlsx"）
            valid_departments: 有效部门集合
            
        Returns:
            ParseResult: 解析结果对象
            
        Raises:
            FileParseError: 不支持的文件格式
        """
        ext = file_extension.lower()
        
        if ext not in self.SUPPORTED_EXTENSIONS:
            raise FileParseError(
                f"不支持的文件格式: {ext}，支持的格式: {', '.join(self.SUPPORTED_EXTENSIONS)}"
            )
        
        if ext == ".csv":
            parser = CsvParser(mode=self.mode, encoding=self.encoding)
        else:
            parser = ExcelParser(mode=self.mode)
        
        return parser.parse(file_content, valid_departments)
    
    def generate_template(self, format: str = "xlsx") -> bytes:
        """
        生成导入模板文件
        
        Args:
            format: 模板格式 ("xlsx" 或 "csv")
            
        Returns:
            bytes: 模板文件二进制内容
        """
        if format.lower() == "csv":
            return self._generate_csv_template()
        elif format.lower() == "xlsx":
            return self._generate_xlsx_template()
        else:
            raise FileParseError(f"不支持的模板格式: {format}，支持 xlsx 和 csv")
    
    def _generate_csv_template(self) -> bytes:
        """
        生成 CSV 格式模板
        
        Returns:
            bytes: CSV 文件二进制内容（UTF-8 BOM）
        """
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        headers = list(self.FIELD_MAPPING.keys())
        writer.writerow(headers)
        
        # 写入示例数据
        example_row = [
            "ASSET001",        # 资产编号
            "测试资产",         # 资产名称
            "EQUIPMENT",       # 资产类型
            "2024-01-01",     # 购置日期
            "10000.00",       # 购置金额
            "DEPT001",        # 使用部门
            "ACTIVE",         # 资产状态
            "办公室A",         # 存放地点
            "这是一条测试数据"  # 备注
        ]
        writer.writerow(example_row)
        
        content = output.getvalue()
        # 添加 UTF-8 BOM
        return "\ufeff".encode("utf-8") + content.encode("utf-8")
    
    def _generate_xlsx_template(self) -> bytes:
        """
        生成 Excel 格式模板
        
        Returns:
            bytes: Excel 文件二进制内容
        """
        if not ExcelParser()._openpyxl_available:
            raise FileParseError("请安装 openpyxl: pip install openpyxl")
        
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "资产导入模板"
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # 写入表头
        headers = list(self.FIELD_MAPPING.keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # 设置列宽
        column_widths = [15, 20, 15, 15, 15, 15, 12, 15, 30]
        for col_idx, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # 写入示例数据
        example_data = [
            ["ASSET001", "测试资产", "EQUIPMENT", "2024-01-01", 10000.00, "DEPT001", "ACTIVE", "办公室A", "这是一条测试数据"]
        ]
        for row_idx, row_data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 添加数据验证说明
        sheet.cell(row=4, column=1, value="资产类型枚举值: EQUIPMENT, INSTRUMENT, VEHICLE, OTHER")
        sheet.cell(row=5, column=1, value="资产状态枚举值: ACTIVE, MAINTENANCE, SCRAPPED")
        sheet.cell(row=6, column=1, value="日期格式: YYYY-MM-DD")
        
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        
        return output.getvalue()


# 导出公共接口
__all__ = [
    "FileParser",
    "CsvParser", 
    "ExcelParser",
    "BaseParser",
    "FileParseError",
    "ValidationError",
    "ParseResult",
    "AssetRowData",
    "AssetTypeEnum",
    "AssetStatusEnum",
]