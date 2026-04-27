"""
资产批量导出 API 模块

支持 CSV/Excel 格式的资产数据导出功能。

功能特性:
- CSV 格式导出
- Excel (.xlsx) 格式导出
- 支持按资产类型、状态、时间范围筛选
- 分页导出支持
- 大数据量内存优化处理

规格来源: SWARM-2025-Q2-P2-006
"""

import io
import csv
from datetime import datetime
from typing import Optional, List, Dict, Any
from decimal import Decimal

from fastapi import APIRouter, Query, HTTPException, status
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel

# 尝试导入 openpyxl，支持 Excel 导出
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORTED = True
except ImportError:
    EXCEL_SUPPORTED = False

from src.models.asset import Asset
from src.repositories.asset_repository import AssetRepository
from src.services.export_service import ExportService


router = APIRouter(prefix="/assets", tags=["资产导出"])


class ExportParams(BaseModel):
    """导出参数模型"""
    asset_type: Optional[str] = None
    status: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    page: int = 1
    page_size: int = 1000


class ExportResponse(BaseModel):
    """导出响应模型"""
    total_count: int
    format: str
    filename: str
    generated_at: str


# 资产导出字段映射（中英文对照）
ASSET_EXPORT_COLUMNS = [
    ("asset_id", "资产编号"),
    ("asset_name", "资产名称"),
    ("asset_type", "资产类型"),
    ("serial_number", "序列号"),
    ("purchase_date", "购置日期"),
    ("purchase_price", "购置价格"),
    ("currency", "币种"),
    ("department", "所属部门"),
    ("custodian", "保管人"),
    ("status", "状态"),
    ("location", "存放地点"),
    ("remarks", "备注"),
]

# 资产类型枚举映射
ASSET_TYPE_ENUM = {
    "EQUIPMENT": "设备",
    "FURNITURE": "家具",
    "VEHICLE": "车辆",
    "IT_HARDWARE": "IT设备",
    "OTHER": "其他",
}

# 资产状态枚举映射
ASSET_STATUS_ENUM = {
    "ACTIVE": "在用",
    "INACTIVE": "闲置",
    "MAINTENANCE": "维保中",
    "RETIRED": "已退役",
}


def format_asset_value(value: Any) -> str:
    """
    格式化资产字段值用于导出
    
    Args:
        value: 字段原始值
        
    Returns:
        str: 格式化后的字符串
    """
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        return str(value)
    return str(value)


def get_export_filename(format: str) -> str:
    """
    生成导出文件名
    
    Args:
        format: 导出格式 (csv/xlsx)
        
    Returns:
        str: 包含时间戳的文件名
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"asset_export_{timestamp}.{format}"


def export_assets_to_csv(
    assets: List[Dict[str, Any]],
    columns: List[tuple]
) -> io.StringIO:
    """
    将资产数据导出为 CSV 格式
    
    Args:
        assets: 资产数据列表
        columns: 列定义 [(字段名, 中文名), ...]
        
    Returns:
        io.StringIO: CSV 文件内容
    """
    output = io.StringIO()
    output.write("\ufeff")  # UTF-8 BOM
    
    writer = csv.writer(output)
    
    # 写入表头
    headers = [col[1] for col in columns]  # 中文表头
    writer.writerow(headers)
    
    # 写入数据行
    for asset in assets:
        row = []
        for field, _ in columns:
            value = asset.get(field, "")
            row.append(format_asset_value(value))
        writer.writerow(row)
    
    output.seek(0)
    return output


def export_assets_to_excel(
    assets: List[Dict[str, Any]],
    columns: List[tuple]
) -> io.BytesIO:
    """
    将资产数据导出为 Excel 格式
    
    Args:
        assets: 资产数据列表
        columns: 列定义 [(字段名, 中文名), ...]
        
    Returns:
        io.BytesIO: Excel 文件内容
        
    Raises:
        HTTPException: 当 Excel 支持不可用时
    """
    if not EXCEL_SUPPORTED:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel export is not supported. Please install openpyxl."
        )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "资产导出"
    
    # 写入表头
    headers = [col[1] for col in columns]
    ws.append(headers)
    
    # 写入数据行
    for asset in assets:
        row = []
        for field, _ in columns:
            value = asset.get(field, "")
            row.append(format_asset_value(value))
        ws.append(row)
    
    # 自动调整列宽
    for col_idx, (_, header) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = len(header)
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(max_length + 5, 50)  # 限制最大宽度
        ws.column_dimensions[col_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/export", response_model=ExportResponse)
async def export_assets(
    format: str = Query("csv", description="导出格式: csv 或 xlsx"),
    asset_type: Optional[str] = Query(None, description="资产类型筛选"),
    status: Optional[str] = Query(None, description="资产状态筛选"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(1000, ge=1, le=10000, description="每页数量")
) -> Response:
    """
    导出资产列表数据
    
    支持 CSV 和 Excel 两种格式，支持筛选和分页。
    
    Args:
        format: 导出格式 (csv/xlsx)
        asset_type: 按资产类型筛选
        status: 按资产状态筛选
        start_date: 开始日期
        end_date: 结束日期
        page: 页码 (默认1)
        page_size: 每页数量 (默认1000，最大10000)
        
    Returns:
        StreamingResponse: 文件流
    """
    # 参数验证
    if format.lower() not in ("csv", "xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Format must be 'csv' or 'xlsx'"
        )
    
    if format.lower() == "xlsx" and not EXCEL_SUPPORTED:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel export is not supported. Please install openpyxl."
        )
    
    # 日期格式验证
    if start_date:
        try:
            datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="start_date must be in YYYY-MM-DD format"
            )
    
    if end_date:
        try:
            datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="end_date must be in YYYY-MM-DD format"
            )
    
    # 构建查询条件
    filters = {}
    if asset_type:
        filters["asset_type"] = asset_type
    if status:
        filters["status"] = status
    if start_date:
        filters["start_date"] = start_date
    if end_date:
        filters["end_date"] = end_date
    
    # 获取导出服务
    export_service = ExportService()
    
    # 获取资产数据（使用流式处理避免内存溢出）
    assets_data = export_service.get_assets_for_export(
        filters=filters,
        page=page,
        page_size=page_size
    )
    
    filename = get_export_filename(format.lower())
    
    if format.lower() == "csv":
        # CSV 导出
        csv_output = export_assets_to_csv(assets_data, ASSET_EXPORT_COLUMNS)
        return StreamingResponse(
            iter([csv_output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    else:
        # Excel 导出
        excel_output = export_assets_to_excel(assets_data, ASSET_EXPORT_COLUMNS)
        return StreamingResponse(
            iter([excel_output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


@router.get("/export/count")
async def get_export_count(
    asset_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
) -> Dict[str, int]:
    """
    获取导出数据总数（用于前端显示）
    
    Returns:
        dict: 符合条件的资产总数
    """
    filters = {}
    if asset_type:
        filters["asset_type"] = asset_type
    if status:
        filters["status"] = status
    if start_date:
        filters["start_date"] = start_date
    if end_date:
        filters["end_date"] = end_date
    
    export_service = ExportService()
    count = export_service.count_assets_for_export(filters)
    
    return {"total": count}


@router.get("/export/preview")
async def preview_export(
    asset_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    limit: int = Query(10, ge=1, le=100, description="预览行数")
) -> Dict[str, Any]:
    """
    导出预览（返回前 N 行数据用于前端预览）
    
    Args:
        limit: 预览行数 (默认10，最大100)
        
    Returns:
        dict: 预览数据及列定义
    """
    export_service = ExportService()
    
    preview_data = export_service.get_assets_for_export(
        filters={"asset_type": asset_type, "status": status} if asset_type or status else {},
        page=1,
        page_size=limit
    )
    
    return {
        "columns": [{"field": col[0], "header": col[1]} for col in ASSET_EXPORT_COLUMNS],
        "rows": preview_data,
        "sample_count": len(preview_data)
    }