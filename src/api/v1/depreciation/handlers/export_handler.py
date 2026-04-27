"""
Depreciation Report Export Handler

Provides CSV and Excel export functionality for depreciation reports.
Implements ATB-009: Report Export CSV verification.
"""

import csv
import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from src.swarm_003.depreciation.domain.schemas import (
    DepreciationDetailResponse,
    DepreciationReportResponse,
)
from src.swarm_003.depreciation.services.depreciation_service import (
    DepreciationService,
)

router = APIRouter(prefix="/api/v1/depreciation", tags=["depreciation-export"])


def get_depreciation_service() -> DepreciationService:
    """Dependency injection for DepreciationService."""
    return DepreciationService()


@router.get(
    "/report/export",
    summary="Export depreciation report",
    description="Export depreciation report in CSV or Excel format",
    responses={
        200: {
            "content": {
                "text/csv": {},
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}
            },
            "description": "Exported file"
        },
        400: {"description": "Invalid parameters"},
        404: {"description": "No data found for the specified period"},
    }
)
async def export_depreciation_report(
    format: str = Query(
        default="csv",
        description="Export format: csv or excel",
        pattern="^(csv|excel)$"
    ),
    period: Optional[str] = Query(
        default=None,
        description="Report period in YYYY-MM format",
        pattern="^\\d{4}-(0[1-9]|1[0-2])$"
    ),
    asset_id: Optional[str] = Query(
        default=None,
        description="Filter by specific asset ID"
    ),
    start_date: Optional[str] = Query(
        default=None,
        description="Start date for report range (YYYY-MM)",
        pattern="^\\d{4}-(0[1-9]|1[0-2])$"
    ),
    end_date: Optional[str] = Query(
        default=None,
        description="End date for report range (YYYY-MM)",
        pattern="^\\d{4}-(0[1-9]|1[0-2])$"
    ),
    service: DepreciationService = Depends(get_depreciation_service)
) -> StreamingResponse:
    """
    Export depreciation report in specified format.
    
    - **format**: Export format (csv or excel)
    - **period**: Single month period (YYYY-MM)
    - **start_date**: Start of date range for multi-month report
    - **end_date**: End of date range for multi-month report
    
    ATB-009 Verification:
    - CSV format returns Content-Type: text/csv
    - Content-Disposition: attachment; filename=depreciation_{period}.csv
    """
    # Validate date range does not exceed 36 months (ATB-010)
    if start_date and end_date:
        start_year, start_month = map(int, start_date.split("-"))
        end_year, end_month = map(int, end_date.split("-"))
        months_diff = (end_year - start_year) * 12 + (end_month - start_month)
        if months_diff > 36:
            raise HTTPException(
                status_code=400,
                detail={
                    "error_code": "PERIOD_EXCEEDS_LIMIT",
                    "message": "Report period cannot exceed 36 months"
                }
            )
    
    # Determine the report period for filename
    report_period = period or start_date or datetime.now().strftime("%Y-%m")
    filename_base = f"depreciation_{report_period}"
    
    # Fetch report data
    try:
        if period:
            report_data = await service.get_aggregate_report(period)
            detail_records = await service.get_depreciation_details(
                asset_id=asset_id,
                start_date=period,
                end_date=period
            )
        elif start_date and end_date:
            report_data = await service.get_aggregate_report(start_date)
            detail_records = await service.get_depreciation_details(
                asset_id=asset_id,
                start_date=start_date,
                end_date=end_date
            )
        elif asset_id:
            current_month = datetime.now().strftime("%Y-%m")
            detail_records = await service.get_depreciation_details(
                asset_id=asset_id,
                start_date=current_month,
                end_date=current_month
            )
            report_data = DepreciationReportResponse(
                period=current_month,
                total_depreciation_amount=sum(
                    r.monthly_depreciation for r in detail_records
                ),
                total_assets=len(detail_records)
            )
        else:
            current_month = datetime.now().strftime("%Y-%m")
            report_data = await service.get_aggregate_report(current_month)
            detail_records = await service.get_depreciation_details(
                start_date=current_month,
                end_date=current_month
            )
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))
    
    # Export based on format
    if format.lower() == "csv":
        return _export_csv(report_period, detail_records, report_data, filename_base)
    elif format.lower() == "excel":
        return _export_excel(report_period, detail_records, report_data, filename_base)
    else:
        raise HTTPException(status_code=400, detail="Unsupported export format")


def _export_csv(
    period: str,
    records: list[DepreciationDetailResponse],
    summary: DepreciationReportResponse,
    filename_base: str
) -> StreamingResponse:
    """
    Generate CSV export of depreciation report.
    
    ATB-009: CSV Export verification
    - Content-Type: text/csv
    - Content-Disposition: attachment; filename=depreciation_{period}.csv
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header with report metadata
    writer.writerow(["Depreciation Report"])
    writer.writerow(["Period", period])
    writer.writerow(["Generated At", datetime.now().isoformat()])
    writer.writerow(["Total Assets", summary.total_assets])
    writer.writerow(["Total Depreciation Amount", str(summary.total_depreciation_amount)])
    writer.writerow([])  # Empty row separator
    
    # Write column headers for detail records
    headers = [
        "Asset ID",
        "Asset Name",
        "Acquisition Date",
        "Acquisition Cost",
        "Useful Life (Months)",
        "Salvage Value",
        "Depreciation Method",
        "Period",
        "Monthly Depreciation",
        "Accumulated Depreciation",
        "Book Value"
    ]
    writer.writerow(headers)
    
    # Write detail records
    for record in records:
        writer.writerow([
            record.asset_id,
            record.asset_name,
            record.acquisition_date.isoformat() if hasattr(record.acquisition_date, 'isoformat') else str(record.acquisition_date),
            str(record.acquisition_cost),
            record.useful_life_months,
            str(record.salvage_value),
            record.depreciation_method.value if hasattr(record.depreciation_method, 'value') else record.depreciation_method,
            record.period,
            str(record.monthly_depreciation),
            str(record.accumulated_depreciation),
            str(record.book_value)
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename_base}.csv"
        }
    )


def _export_excel(
    period: str,
    records: list[DepreciationDetailResponse],
    summary: DepreciationReportResponse,
    filename_base: str
) -> StreamingResponse:
    """
    Generate Excel export of depreciation report.
    
    Returns:
        Excel file (.xlsx) with:
        - Sheet 1: Summary
        - Sheet 2: Detail Records
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return _export_csv(period, records, summary, filename_base)
    
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    
    # Sheet 1: Summary
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    
    summary_sheet.append(["Depreciation Report Summary"])
    summary_sheet.append(["Period", period])
    summary_sheet.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_sheet.append(["Total Assets", summary.total_assets])
    summary_sheet.append(["Total Depreciation Amount", float(summary.total_depreciation_amount)])
    
    # Style summary header
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet.column_dimensions["A"].width = 25
    summary_sheet.column_dimensions["B"].width = 20
    
    # Sheet 2: Detail Records
    detail_sheet = workbook.create_sheet("Detail Records")
    
    # Headers
    detail_headers = [
        "Asset ID",
        "Asset Name",
        "Acquisition Date",
        "Acquisition Cost",
        "Useful Life (Months)",
        "Salvage Value",
        "Depreciation Method",
        "Period",
        "Monthly Depreciation",
        "Accumulated Depreciation",
        "Book Value"
    ]
    detail_sheet.append(detail_headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in detail_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Write detail records
    for record in records:
        detail_sheet.append([
            record.asset_id,
            record.asset_name,
            str(record.acquisition_date),
            float(record.acquisition_cost),
            record.useful_life_months,
            float(record.salvage_value),
            record.depreciation_method.value if hasattr(record.depreciation_method, 'value') else record.depreciation_method,
            record.period,
            float(record.monthly_depreciation),
            float(record.accumulated_depreciation),
            float(record.book_value)
        ])
    
    # Adjust column widths
    for col_idx, width in enumerate([20, 25, 15, 15, 15, 15, 20, 12, 18, 20, 15], 1):
        detail_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    workbook.save(output)
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename_base}.xlsx"
        }
    )


@router.get(
    "/report/export/validate",
    summary="Validate export parameters",
    description="Validate export parameters without generating file"
)
async def validate_export_params(
    start_date: str = Query(description="Start date (YYYY-MM)"),
    end_date: str = Query(description="End date (YYYY-MM)"),
) -> dict:
    """
    Validate export parameters before triggering export.
    
    Returns validation result including period span in months.
    """
    start_year, start_month = map(int, start_date.split("-"))
    end_year, end_month = map(int, end_date.split("-"))
    months_diff = (end_year - start_year) * 12 + (end_month - start_month)
    
    return {
        "valid": months_diff <= 36,
        "period_span_months": months_diff,
        "max_allowed_months": 36,
        "error" if months_diff > 36 else None: (
            "Period exceeds maximum allowed range of 36 months"
            if months_diff > 36 else None
        )
    }